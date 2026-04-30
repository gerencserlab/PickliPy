from __future__ import annotations

import copy
import random
import secrets
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import openpyxl
from openpyxl.utils.cell import absolute_coordinate, get_column_letter

import warnings

warnings.filterwarnings(
    "ignore",
    message="Data Validation extension is not supported",
)

try:
    from .common import (
        PLATE_COLS_384,
        PLATE_ROWS_384,
        PicklyPyConfigError,
        PicklyPyError,
        as_str,
        run_with_user_facing_errors,
    )
except ImportError:  # Allows direct execution from the source folder during development.
    from common import (  # type: ignore
        PLATE_COLS_384,
        PLATE_ROWS_384,
        PicklyPyConfigError,
        PicklyPyError,
        as_str,
        run_with_user_facing_errors,
    )


VERSION = "0.1-python"

DEFAULT_PLATE_MAP_LABEL = "Plate Map:"
DEFAULT_SCRAMBLED_LABEL = "Scrambled Plate Map:"
DEFAULT_SOURCE_START_COL = 2  # B. PicklyPy plate maps begin one cell right of the label.
DEFAULT_EDGE_WIDTH = 2


GridPos = Tuple[int, int]  # 0-based row, 0-based column within the plate map.


@dataclass(frozen=True)
class ScrambleSummary:
    plate_map_row: int
    source_cells: int
    target_start_cell: str


def _print_banner() -> None:
    print("PicklyPy.scramble: add constrained scrambled plate-map formula tables.")
    print(f"Python version {VERSION}")


def _is_blank_value(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    return False


def _display_value(value: Any) -> str:
    """String conversion used only for detecting Plate Map labels in column A."""
    return as_str(value).strip()


def _get_dst_sheet(
    wb: openpyxl.Workbook,
    dst_sheet: Optional[str],
) -> openpyxl.worksheet.worksheet.Worksheet:
    """Return the DST worksheet.

    PicklyPy.Assay reads the second worksheet as DST.  This helper keeps that
    behavior unless the user explicitly supplies --dst-sheet.
    """
    if dst_sheet is not None:
        if dst_sheet not in wb.sheetnames:
            raise PicklyPyConfigError(f"DST worksheet not found: {dst_sheet!r}")
        return wb[dst_sheet]

    if "DST" in wb.sheetnames:
        return wb["DST"]

    if len(wb.worksheets) < 2:
        raise PicklyPyConfigError(
            "DST worksheet not found. Supply --dst-sheet, or use a workbook with a second worksheet."
        )

    return wb.worksheets[1]


def _find_plate_map_rows(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    *,
    label: str,
) -> List[int]:
    """Find 1-based worksheet rows whose first cell exactly matches label."""
    rows: List[int] = []
    max_row = ws.max_row or 1
    for row in range(1, max_row + 1):
        if _display_value(ws.cell(row=row, column=1).value) == label:
            rows.append(row)
    return rows


def _occupied_positions(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    *,
    label_row: int,
    rows: int,
    cols: int,
    source_start_col: int,
) -> List[GridPos]:
    """Return occupied plate-map cells as 0-based grid positions."""
    out: List[GridPos] = []
    start_row = label_row + 1
    for r in range(rows):
        for c in range(cols):
            value = ws.cell(row=start_row + r, column=source_start_col + c).value
            if not _is_blank_value(value):
                out.append((r, c))
    return out


def _all_plate_positions(rows: int, cols: int) -> List[GridPos]:
    return [(r, c) for r in range(rows) for c in range(cols)]


def _is_edge_position(pos: GridPos, *, rows: int, cols: int, edge_width: int) -> bool:
    if edge_width <= 0:
        return False
    r, c = pos
    return (
        r < edge_width
        or r >= rows - edge_width
        or c < edge_width
        or c >= cols - edge_width
    )


def _allowed_target(
    source: GridPos,
    target: GridPos,
    *,
    rows: int,
    cols: int,
    edge_width: int,
) -> bool:
    """Return True when target satisfies the scrambling constraints.

    Constraints:
      - source and target are not the same well;
      - target is in a different row;
      - target is in a different column;
      - if source is in any edge band, target must be outside all edge bands.

    The edge-band rule treats the outside `edge_width` rows and columns as the
    plate edge.  Sources from that region are therefore relocated into the
    interior rectangle.
    """
    sr, sc = source
    tr, tc = target

    if sr == tr or sc == tc:
        return False

    if edge_width <= 0:
        return True

    if _is_edge_position(source, rows=rows, cols=cols, edge_width=edge_width) and _is_edge_position(
        target, rows=rows, cols=cols, edge_width=edge_width
    ):
        return False

    return True


def _pos_to_a1(
    pos: GridPos,
    *,
    label_row: int,
    start_col: int,
) -> str:
    r, c = pos
    return f"{get_column_letter(start_col + c)}{label_row + 1 + r}"


def _find_random_matching(
    sources: Sequence[GridPos],
    targets: Sequence[GridPos],
    *,
    rows: int,
    cols: int,
    edge_width: int,
    rng: random.Random,
) -> Dict[GridPos, GridPos]:
    """Find a random injective mapping from source positions to target positions.

    This uses a randomized Kuhn augmenting-path matching.  With the default
    occupied-targets mode, len(sources) == len(targets), so the result is a true
    permutation of the occupied plate-map positions.  With --allow-empty-targets,
    targets can be larger than sources, and the result is an injective relocation
    into the full plate.
    """
    if len(sources) == 0:
        return {}

    if len(targets) < len(sources):
        raise PicklyPyConfigError(
            f"Not enough target wells for constrained scramble: {len(sources)} sources, {len(targets)} targets."
        )

    adjacency: List[List[int]] = []
    for source in sources:
        candidates = [
            i
            for i, target in enumerate(targets)
            if _allowed_target(source, target, rows=rows, cols=cols, edge_width=edge_width)
        ]
        rng.shuffle(candidates)
        adjacency.append(candidates)

    no_candidates = [sources[i] for i, candidates in enumerate(adjacency) if not candidates]
    if no_candidates:
        example = no_candidates[0]
        raise PicklyPyConfigError(
            "At least one occupied well has no legal scrambled target. "
            f"First failing source well is grid row {example[0] + 1}, column {example[1] + 1}."
        )

    order = list(range(len(sources)))
    rng.shuffle(order)
    order.sort(key=lambda i: len(adjacency[i]))

    pair_target_to_source = [-1] * len(targets)
    pair_source_to_target = [-1] * len(sources)

    def augment(source_index: int, seen_targets: List[bool]) -> bool:
        for target_index in adjacency[source_index]:
            if seen_targets[target_index]:
                continue
            seen_targets[target_index] = True

            previous_source = pair_target_to_source[target_index]
            if previous_source == -1 or augment(previous_source, seen_targets):
                pair_target_to_source[target_index] = source_index
                pair_source_to_target[source_index] = target_index
                return True

        return False

    for source_index in order:
        seen = [False] * len(targets)
        if not augment(source_index, seen):
            source_pos = sources[source_index]
            raise PicklyPyConfigError(
                "Could not construct a complete constrained scramble for this plate map. "
                f"Failed at grid row {source_pos[0] + 1}, column {source_pos[1] + 1}. "
                "The occupied-well pattern may be too sparse for the row/column/edge constraints."
            )

    if any(target_index < 0 for target_index in pair_source_to_target):
        raise PicklyPyConfigError(
            "Internal error: matching did not assign every occupied source position."
        )

    return {
        source_pos: targets[pair_source_to_target[source_index]]
        for source_index, source_pos in enumerate(sources)
    }


def _target_area_has_content(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    *,
    label_row: int,
    rows: int,
    cols: int,
    target_start_col: int,
) -> List[str]:
    """Return A1 addresses of non-empty cells in the output header/grid area."""
    nonempty: List[str] = []

    # Header/metadata row above the table.
    for c in range(cols):
        cell = ws.cell(row=label_row, column=target_start_col + c)
        if not _is_blank_value(cell.value):
            nonempty.append(cell.coordinate)

    # Formula table.
    for r in range(rows):
        for c in range(cols):
            cell = ws.cell(row=label_row + 1 + r, column=target_start_col + c)
            if not _is_blank_value(cell.value):
                nonempty.append(cell.coordinate)

    return nonempty


def _clear_target_area(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    *,
    label_row: int,
    rows: int,
    cols: int,
    target_start_col: int,
) -> None:
    # Header/metadata row.
    for c in range(cols):
        ws.cell(row=label_row, column=target_start_col + c).value = None

    # Formula table.
    for r in range(rows):
        for c in range(cols):
            ws.cell(row=label_row + 1 + r, column=target_start_col + c).value = None


def _copy_cell_style(
    source: openpyxl.cell.cell.Cell,
    target: openpyxl.cell.cell.Cell,
) -> None:
    """Copy visual cell formatting without copying the cell value."""
    if source.has_style:
        target._style = copy.copy(source._style)
    if source.number_format:
        target.number_format = source.number_format
    if source.alignment:
        target.alignment = copy.copy(source.alignment)
    if source.protection:
        target.protection = copy.copy(source.protection)
    if source.comment:
        target.comment = copy.copy(source.comment)


def _copy_column_widths(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    *,
    source_start_col: int,
    target_start_col: int,
    cols: int,
) -> None:
    for c in range(cols):
        source_letter = get_column_letter(source_start_col + c)
        target_letter = get_column_letter(target_start_col + c)
        source_dim = ws.column_dimensions[source_letter]
        target_dim = ws.column_dimensions[target_letter]
        target_dim.width = source_dim.width


def _write_scrambled_table(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    *,
    label_row: int,
    mapping: Dict[GridPos, GridPos],
    rows: int,
    cols: int,
    source_start_col: int,
    target_start_col: int,
    scrambled_label: str,
    seed: int,
    copy_styles: bool,
) -> None:
    """Write formulas into the target table.

    `mapping` is source -> target, but formulas are written into target cells.
    """
    _clear_target_area(
        ws,
        label_row=label_row,
        rows=rows,
        cols=cols,
        target_start_col=target_start_col,
    )

    # Header/metadata row is intentionally outside the formula table.
    label_cell = ws.cell(row=label_row, column=target_start_col)
    label_cell.value = scrambled_label
    _copy_cell_style(ws.cell(row=label_row, column=1), label_cell)

    seed_cell = ws.cell(row=label_row, column=target_start_col + 1)
    seed_cell.value = f"Seed: {seed}"
    _copy_cell_style(ws.cell(row=label_row, column=2), seed_cell)

    for source_pos, target_pos in mapping.items():
        source_cell = ws.cell(
            row=label_row + 1 + source_pos[0],
            column=source_start_col + source_pos[1],
        )
        target_cell = ws.cell(
            row=label_row + 1 + target_pos[0],
            column=target_start_col + target_pos[1],
        )

        if copy_styles:
            _copy_cell_style(source_cell, target_cell)

        #target_cell.value = f"={absolute_coordinate(source_cell.coordinate)}"
        target_cell.value = f"={source_cell.coordinate}"

    _copy_column_widths(
        ws,
        source_start_col=source_start_col,
        target_start_col=target_start_col,
        cols=cols,
    )


def scramble_workbook(
    xlsx_path: str | Path,
    *,
    output: Optional[str | Path] = None,
    in_place: bool = False,
    dst_sheet: Optional[str] = None,
    seed: Optional[int] = None,
    force: bool = False,
    allow_empty_targets: bool = False,
    rows: int = PLATE_ROWS_384,
    cols: int = PLATE_COLS_384,
    edge_width: int = DEFAULT_EDGE_WIDTH,
    source_start_col: int = DEFAULT_SOURCE_START_COL,
    plate_map_label: str = DEFAULT_PLATE_MAP_LABEL,
    scrambled_label: str = DEFAULT_SCRAMBLED_LABEL,
    copy_styles: bool = True,
) -> List[ScrambleSummary]:
    """Add scrambled formula tables to every Plate Map in the DST worksheet.

    The default target set is the set of occupied cells in the original plate map.
    That preserves the original empty-cell mask while still permuting every occupied
    well to a different row and different column.  Use --allow-empty-targets to
    instead relocate occupied wells into any legal well of the plate.
    """
    _print_banner()

    input_path = Path(xlsx_path)
    if not input_path.exists():
        raise PicklyPyConfigError(f"Workbook not found: {input_path}")

    if in_place and output is not None:
        raise PicklyPyConfigError("Use either --in-place or --output, not both.")

    if seed is None:
        seed = secrets.randbits(63)
    rng = random.Random(seed)

    if rows <= 0 or cols <= 0:
        raise PicklyPyConfigError("--rows and --cols must be positive integers.")
    if edge_width < 0:
        raise PicklyPyConfigError("--edge-width must be zero or a positive integer.")
    if edge_width * 2 >= rows or edge_width * 2 >= cols:
        raise PicklyPyConfigError(
            "--edge-width is too large for the requested plate dimensions."
        )

    target_start_col = source_start_col + cols

    wb = openpyxl.load_workbook(input_path, data_only=False)
    ws = _get_dst_sheet(wb, dst_sheet)

    plate_map_rows = _find_plate_map_rows(ws, label=plate_map_label)
    if not plate_map_rows:
        raise PicklyPyConfigError(
            f"No {plate_map_label!r} section found in worksheet {ws.title!r}."
        )

    summaries: List[ScrambleSummary] = []

    print(f"Workbook: {input_path}")
    print(f"DST worksheet: {ws.title}")
    print(f"Random seed: {seed}")

    for label_row in plate_map_rows:
        occupied = _occupied_positions(
            ws,
            label_row=label_row,
            rows=rows,
            cols=cols,
            source_start_col=source_start_col,
        )

        target_start_cell = f"{get_column_letter(target_start_col)}{label_row + 1}"

        existing = _target_area_has_content(
            ws,
            label_row=label_row,
            rows=rows,
            cols=cols,
            target_start_col=target_start_col,
        )
        if existing and not force:
            preview = ", ".join(existing[:8])
            more = "" if len(existing) <= 8 else f", ... ({len(existing)} cells total)"
            raise PicklyPyConfigError(
                "The target scrambled table area already contains content at "
                f"{preview}{more}. Re-run with --force to overwrite it."
            )

        if len(occupied) == 0:
            _write_scrambled_table(
                ws,
                label_row=label_row,
                mapping={},
                rows=rows,
                cols=cols,
                source_start_col=source_start_col,
                target_start_col=target_start_col,
                scrambled_label=scrambled_label,
                seed=seed,
                copy_styles=copy_styles,
            )
            print(
                f"Plate Map row {label_row}: original map is empty; wrote empty scrambled table at {target_start_cell}."
            )
            summaries.append(
                ScrambleSummary(
                    plate_map_row=label_row,
                    source_cells=0,
                    target_start_cell=target_start_cell,
                )
            )
            continue

        if len(occupied) == 1:
            only = occupied[0]
            raise PicklyPyConfigError(
                "Cannot scramble a plate map with only one occupied well because the target must be "
                f"in a different row and different column. First occupied well: {_pos_to_a1(only, label_row=label_row, start_col=source_start_col)}."
            )

        targets = _all_plate_positions(rows, cols) if allow_empty_targets else list(occupied)

        mapping = _find_random_matching(
            occupied,
            targets,
            rows=rows,
            cols=cols,
            edge_width=edge_width,
            rng=rng,
        )

        _write_scrambled_table(
            ws,
            label_row=label_row,
            mapping=mapping,
            rows=rows,
            cols=cols,
            source_start_col=source_start_col,
            target_start_col=target_start_col,
            scrambled_label=scrambled_label,
            seed=seed,
            copy_styles=copy_styles,
        )

        edge_sources = sum(
            1
            for source in occupied
            if _is_edge_position(source, rows=rows, cols=cols, edge_width=edge_width)
        )
        print(
            f"Plate Map row {label_row}: scrambled {len(occupied)} occupied wells "
            f"({edge_sources} from edge bands) into {target_start_cell}."
        )

        summaries.append(
            ScrambleSummary(
                plate_map_row=label_row,
                source_cells=len(occupied),
                target_start_cell=target_start_cell,
            )
        )

    if in_place:
        output_path = input_path
    elif output is not None:
        output_path = Path(output)
    else:
        output_path = input_path.with_name(f"{input_path.stem}_scrambled{input_path.suffix}")

    wb.save(output_path)
    print(f"Saved: {output_path}")

    return summaries


def main(argv: Optional[Sequence[str]] = None) -> int:
    import argparse

    parser = argparse.ArgumentParser(
        prog="PicklyPy.scramble",
        description=(
            "Add a constrained scrambled plate-map formula table to every Plate Map "
            "in the DST worksheet of a PicklyPy assay workbook."
        ),
    )
    parser.add_argument("xlsx", help="Path to the PicklyPy assay .xlsx workbook")
    parser.add_argument(
        "-o",
        "--output",
        help="Output workbook path. Defaults to '<input>_scrambled.xlsx' unless --in-place is used.",
    )
    parser.add_argument(
        "--in-place",
        action="store_true",
        help="Modify the input workbook in place instead of writing a '<input>_scrambled.xlsx' copy.",
    )
    parser.add_argument(
        "--dst-sheet",
        help="DST worksheet name. Defaults to sheet named 'DST', otherwise the second worksheet like PicklyPy.Assay.",
    )
    parser.add_argument(
        "--seed",
        type=int,
        help="Integer random seed for reproducible scrambling. If omitted, a seed is generated and printed.",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Overwrite any existing content in the scrambled table area.",
    )
    parser.add_argument(
        "--allow-empty-targets",
        action="store_true",
        help=(
            "Allow occupied wells to move into any legal plate position. "
            "Default is to target only originally occupied positions, preserving the original empty-cell mask."
        ),
    )
    parser.add_argument(
        "--edge-width",
        type=int,
        default=DEFAULT_EDGE_WIDTH,
        help="Number of outside rows/columns treated as plate edge bands. Default: 2.",
    )
    parser.add_argument(
        "--rows",
        type=int,
        default=PLATE_ROWS_384,
        help=f"Plate-map row count. Default: {PLATE_ROWS_384}.",
    )
    parser.add_argument(
        "--cols",
        type=int,
        default=PLATE_COLS_384,
        help=f"Plate-map column count. Default: {PLATE_COLS_384}.",
    )
    parser.add_argument(
        "--source-start-col",
        type=int,
        default=DEFAULT_SOURCE_START_COL,
        help="1-based column where the original plate-map grid starts. Default: 2 (B).",
    )
    parser.add_argument(
        "--plate-map-label",
        default=DEFAULT_PLATE_MAP_LABEL,
        help=f"Column-A label identifying plate maps. Default: {DEFAULT_PLATE_MAP_LABEL!r}.",
    )
    parser.add_argument(
        "--scrambled-label",
        default=DEFAULT_SCRAMBLED_LABEL,
        help=f"Header label to write above each scrambled table. Default: {DEFAULT_SCRAMBLED_LABEL!r}.",
    )
    parser.add_argument(
        "--no-copy-styles",
        action="store_true",
        help="Do not copy cell styles from source wells to their scrambled formula cells.",
    )
    parser.add_argument(
        "--no-pause",
        action="store_true",
        help="Do not wait for Enter at the end.",
    )

    args = parser.parse_args(list(argv) if argv is not None else None)

    return run_with_user_facing_errors(
        lambda: scramble_workbook(
            args.xlsx,
            output=args.output,
            in_place=args.in_place,
            dst_sheet=args.dst_sheet,
            seed=args.seed,
            force=args.force,
            allow_empty_targets=args.allow_empty_targets,
            rows=args.rows,
            cols=args.cols,
            edge_width=args.edge_width,
            source_start_col=args.source_start_col,
            plate_map_label=args.plate_map_label,
            scrambled_label=args.scrambled_label,
            copy_styles=not args.no_copy_styles,
        ),
        pause_on_exit=not args.no_pause,
    )


if __name__ == "__main__":
    raise SystemExit(main())
