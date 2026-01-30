from __future__ import annotations

import copy
import math
import os
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Optional, Sequence, Tuple

import openpyxl

import warnings
warnings.filterwarnings(
    "ignore",
    message="Data Validation extension is not supported",
)

from .common import (
    PLATE_COLS_384,
    PLATE_ROWS_384,
    PicklyPyConfigError,
    PicklyPyError,
    PicklyPyVolumeError,
    as_number,
    as_str,
    format_csv_rows,
    integer_chop,
    load_survey_volumes,
    load_workbook,
    make_inventory_row,
    parse_well_name,
    quote_csv_field,
    reorder_destinations_within_group,
    reorder_source_groups,
    round_to_increment,
    run_with_user_facing_errors,
    sheet_to_matrix,
    split_labels,
    write_text,
    well_name
)

from .randomization import (
    apply_well_map_to_str_grid,
    apply_well_map_to_well_dict,
    optimize_plate_well_map,
    seed_from_text,
)


VERSION = "4.19-python"
SOURCE_WLS_VERSION = "4.19  01/29/2026"


@dataclass
class AssayHeader:
    inventory_src: str = "Inventory"
    inventory_dst: str = "Inventory"
    rack_src: int = 1
    rack_dst: int = 2
    process_src: str = "Process_SRC"
    process_dst: str = "Process_DST"
    well_volume_ul_default: float = 0.0


@dataclass
class Addition:
    index: int
    picklist_name: str
    barcode_src: str
    barcodes_dst: List[str]
    well_volume_ul: float
    plate_map_row: int  # 0-based row index of the 'Plate Map:' label
    label_row: int  # 0-based row index of the 'Labels' label
    label_rows: List[List[Any]]  # rows from the Labels list, cols A-C
    conc_map_row: Optional[int] = None  # 0-based row index of 'Concentrations Map:' label


def _print_banner() -> None:
    print(
        "Picklist generator for acoustic dispensing using Beckman Coulter Echo 650 "
        "controlled by Revvity PlateWorks. Written by Akos A. Gerencser."
    )
    print(f"Python translation version {VERSION} (from WL {SOURCE_WLS_VERSION})")
    print(
        "Supporting multiple source and destination plates, volume checking and distributing "
        "larger total volume additions into multiple wells."
    )
    print(
        "Supporting dose-response, use 'Concentrations Map:' label below the plate map to provide concentrations. "
        "For dose response, only one compound per well is supported in one table; dispense multiple compounds using "
        "a new plate map with the same picklist name."
    )
    print(
        "Creates inventory and process files."
        "Barcodes will come alphabetically."
    )


def _find_rows_with_label(dst: List[List[Any]], label: str) -> List[int]:
    out: List[int] = []
    for i, row in enumerate(dst):
        if len(row) == 0:
            continue
        if as_str(row[0]) == label:
            out.append(i)
    return out


def _extract_plate_grid(dst: List[List[Any]], label_row: int) -> List[List[str]]:
    """Extract a 16x24 plate grid starting one cell down/right from label_row."""
    start_r = label_row + 1
    end_r = start_r + PLATE_ROWS_384
    if end_r > len(dst):
        raise PicklyPyConfigError(f"DST worksheet does not contain a full {PLATE_ROWS_384}-row plate map.")
    grid: List[List[str]] = []
    for r in range(start_r, end_r):
        row = dst[r]
        # Columns B..Y => indices 1..24 (inclusive end 25)
        cells = [as_str(v) for v in row[1 : 1 + PLATE_COLS_384]]
        # Pad if needed
        if len(cells) < PLATE_COLS_384:
            cells += [""] * (PLATE_COLS_384 - len(cells))
        grid.append(cells)
    return grid


def _is_grid_empty(grid: List[List[str]]) -> bool:
    return all(cell.strip() == "" for row in grid for cell in row)


def _parse_header(dst: List[List[Any]]) -> AssayHeader:
    # WL uses first 8 rows, columns A and B.
    header: Dict[str, Any] = {}
    for r in range(min(8, len(dst))):
        key = as_str(dst[r][0])
        val = dst[r][1] if len(dst[r]) > 1 else ""
        if key != "":
            header[key] = val

    inv_src = as_str(header.get("Inventory_SRC:", "Inventory")) or "Inventory"
    inv_dst = as_str(header.get("Inventory_DST:", "Inventory")) or "Inventory"
    rack_src = integer_chop(as_number(header.get("Inventory_SRC_Rack#:", 1)) or 1)
    rack_dst = integer_chop(as_number(header.get("Inventory_DST_Rack#:", 2)) or 2)
    proc_src = as_str(header.get("Process_SRC:", "Process_SRC")) or "Process_SRC"
    proc_dst = as_str(header.get("Process_DST:", "Process_DST")) or "Process_DST"

    wv = as_number(header.get("Well volume (ul):", None))
    if wv is None:
        # WL falls back to dst[[1,2]] i.e. B1
        try:
            wv = as_number(dst[0][1])
        except Exception:
            wv = None

    if wv is None or not math.isfinite(float(wv)):
        raise PicklyPyConfigError(
            "Error: missing definition of assay (DST) well volume 'Well volume (ul):'"
        )

    return AssayHeader(
        inventory_src=inv_src,
        inventory_dst=inv_dst,
        rack_src=int(rack_src) if isinstance(rack_src, (int,)) else int(float(rack_src)),
        rack_dst=int(rack_dst) if isinstance(rack_dst, (int,)) else int(float(rack_dst)),
        process_src=proc_src,
        process_dst=proc_dst,
        well_volume_ul_default=float(wv),
    )


def _parse_additions(dst: List[List[Any]], well_volume_default: float) -> Tuple[List[Addition], List[str]]:
    plate_map_rows = _find_rows_with_label(dst, "Plate Map:")
    if not plate_map_rows:
        raise PicklyPyConfigError("No 'Plate Map:' section found in DST worksheet.")

    label_rows = _find_rows_with_label(dst, "Labels")
    if not label_rows:
        raise PicklyPyConfigError("No 'Labels' section found in DST worksheet.")

    # Destination barcode universe for '*' expansion:
    all_dst_barcodes_raw: List[str] = []

    additions: List[Addition] = []

    for add_idx, pm_row in enumerate(sorted(plate_map_rows)):
        # Find the Labels row for this plate map: first 'Labels' after pm_row, before next plate map.
        next_pm_row = min([r for r in plate_map_rows if r > pm_row], default=len(dst))
        labels_row = min([r for r in label_rows if pm_row < r < next_pm_row], default=None)
        if labels_row is None:
            raise PicklyPyConfigError(
                "Some definitions are missing. Each plate map must start with 'Plate Map:' and followed by lines defining:\n"
                "Picklist Name:\nBarcode_SRC:\nBarcode_DST:\nLabels"
            )

        # Parameters live between end of plate map grid and the Labels row.
        # WL effectively reads a big A->B mapping and then looks up known keys.
        params: Dict[str, Any] = {}
        for r in range(pm_row + 1, labels_row):
            key = as_str(dst[r][0]) if len(dst[r]) > 0 else ""
            val = dst[r][1] if len(dst[r]) > 1 else ""
            if key != "":
                params[key] = val

        picklist_name = as_str(params.get("Picklist Name:", ""))
        if picklist_name == "":
            raise PicklyPyConfigError(
                "Some definitions are missing. Each plate map must start with 'Plate Map:' and followed by lines defining:\n"
                "Picklist Name:\nBarcode_SRC:\nBarcode_DST:\nLabels"
            )

        barcode_src = as_str(params.get("Barcode_SRC:", "PlateSRC")) or "PlateSRC"
        barcode_dst_raw = as_str(params.get("Barcode_DST:", "PlateDST")) or "PlateDST"
        barcodes_dst = [b.strip() for b in barcode_dst_raw.split(",") if b.strip() != ""]
        if not barcodes_dst:
            barcodes_dst = ["PlateDST"]

        # Optional per-addition well volume override.
        wv = as_number(params.get("Well volume (ul):", None))
        if wv is None:
            wv = well_volume_default

        if wv is None or not math.isfinite(float(wv)):
            raise PicklyPyConfigError(
                "Error: missing definition of assay (DST) well volume 'Well volume (ul):'"
            )

        # Optional concentrations map (dose response) directly below the plate map.
        conc_map_row: Optional[int] = None
        for r in range(pm_row + 1, labels_row):
            lab = as_str(dst[r][0])
            if lab in ("Concentrations Map:", "Concentrations  Map:"):
                conc_map_row = r
                break

        # Label list: starts at labels_row+1, continues until blank in col A or next Plate Map:
        label_list_rows: List[List[Any]] = []
        r = labels_row + 1
        while r < len(dst):
            col_a = as_str(dst[r][0]) if len(dst[r]) > 0 else ""
            if col_a == "" or col_a == "Plate Map:":
                break
            # Take cols A-C
            row_vals = [dst[r][c] if c < len(dst[r]) else "" for c in range(3)]
            label_list_rows.append(row_vals)
            r += 1

        additions.append(
            Addition(
                index=add_idx,
                picklist_name=picklist_name,
                barcode_src=barcode_src,
                barcodes_dst=barcodes_dst,
                well_volume_ul=float(wv),
                plate_map_row=pm_row,
                label_row=labels_row,
                label_rows=label_list_rows,
                conc_map_row=conc_map_row,
            )
        )

        all_dst_barcodes_raw.extend([b for b in barcodes_dst if b != "*"])

    all_dst_barcodes = sorted(set(all_dst_barcodes_raw))
    if not all_dst_barcodes:
        raise PicklyPyConfigError(
            "Barcode_DST definition is missing. Cannot use * for all barcodes."
        )

    return additions, all_dst_barcodes


def _validate_src_sheet(src: List[List[Any]]) -> None:
    # Ensure at least 7 columns so we can access G1 and the 'new' override column.
    for r in range(len(src)):
        if len(src[r]) < 7:
            src[r] = src[r] + [""] * (7 - len(src[r]))

    # Unique (plate barcode, source well)
    keys: List[str] = []
    for row in src[1:]:
        plate = as_str(row[5])
        well = as_str(row[0])
        if plate.strip() == "" or well.strip() == "":
            continue
        keys.append(f"{plate}{well}")

    dup = sorted({k for k in keys if keys.count(k) > 1})
    if dup:
        raise PicklyPyConfigError(
            "Error: duplicate well assignement in 'src' worksheet within the same source plate: "
            + ", ".join(dup)
        )

    # Duplicate compound names within a plate must have the same stock concentration.
    stock_by_key: Dict[Tuple[str, str], set] = {}
    for row in src[1:]:
        compound = as_str(row[1]).strip()
        if compound == "":
            continue
        plate = as_str(row[5]).strip()
        stock = as_number(row[2])
        if stock is None:
            continue
        stock_by_key.setdefault((plate, compound), set()).add(float(stock))

    bad = [(plate, comp, vals) for (plate, comp), vals in stock_by_key.items() if len(vals) > 1]
    if bad:
        raise PicklyPyConfigError(
            "Error: duplicate compound labels in 'src' worksheet must have the same concentrations within each source plate!"
        )


def _apply_survey_to_src(
    src: List[List[Any]],
    survey_dir: Path,
    used_source_barcodes: Optional[Sequence[str]] = None,
) -> None:
    """Update src volumes (col E) from any *_platesurvey.xml files in the folder."""
    survey = load_survey_volumes(survey_dir)
    if not survey.volumes_ul:
        return

    # Some Echo exports use a placeholder barcode "UnknownBarCode". Like the
    # original Wolfram scripts, we only support this if the run uses a single
    # source plate barcode (otherwise the mapping is ambiguous).
    unknown_mapped_to: Optional[str] = None
    has_unknown = any(bc == "UnknownBarCode" for (bc, _w) in survey.volumes_ul)
    if has_unknown:
        if used_source_barcodes is not None:
            used = [as_str(b).strip() for b in used_source_barcodes if as_str(b).strip() not in {"", "*"}]
        else:
            used = [as_str(r[5]).strip() for r in src[1:] if len(r) > 5 and as_str(r[5]).strip() not in {"", "*"}]
        used_unique = sorted(set(used))
        if len(used_unique) > 1:
            raise PicklyPyConfigError(
                "Error: plate survey contains the barcode 'UnknownBarCode', but multiple source plate barcodes are used in the run: "
                + ",".join(used_unique)
                + ". Please re-export the survey with correct barcodes, or run with a single source plate barcode."
            )
        if len(used_unique) == 1:
            unknown_mapped_to = used_unique[0]
            print(
                f"Survey barcode is UnknownBarCode; mapping survey volumes to source plate barcode '{unknown_mapped_to}'."
            )

    # Snapshot original volumes for 'new' override.
    src_orig_vols = [row[4] if len(row) > 4 else "" for row in src]

    # Apply.
    missing_any = False
    for i in range(1, len(src)):
        row = src[i]
        if len(row) < 7:
            row.extend([""] * (7 - len(row)))
        plate = as_str(row[5]).strip()
        well = as_str(row[0]).strip()
        if plate == "" or well == "":
            continue
        vol = survey.get(plate, well)
        if vol is None and unknown_mapped_to == plate:
            vol = survey.get("UnknownBarCode", well)
        if vol is None:
            missing_any = True
            vol = 0.0
        row[4] = float(vol)

    if missing_any:
        # Match WL: warn and use 0.
        print(
            "Using survey file(s) but not finding values for one or more wells; using 0 for those."
        )

    # New addition override: rows with "new" in column G keep original volume.
    for i in range(1, len(src)):
        row = src[i]
        flag = as_str(row[6]).strip()
        if flag.lower() == "new":
            row[4] = src_orig_vols[i]

    print("Volumes updated from survey file(s).")


def _build_srcr(src: List[List[Any]]) -> Dict[str, Dict[str, Tuple[List[str], float]]]:
    """Build mapping: plate_barcode -> compound_name -> (source_wells_list, stock_conc)."""
    by_plate: Dict[str, Dict[str, List[Tuple[str, float]]]] = {}
    for row in src[1:]:
        plate = as_str(row[5]).strip()
        well = as_str(row[0]).strip()
        comp = as_str(row[1]).strip()
        stock = as_number(row[2])
        if plate == "" or well == "" or comp == "" or stock is None:
            continue
        by_plate.setdefault(plate, {}).setdefault(comp, []).append((well, float(stock)))

    srcr: Dict[str, Dict[str, Tuple[List[str], float]]] = {}
    for plate, comp_map in by_plate.items():
        srcr[plate] = {}
        for comp, entries in comp_map.items():
            wells = [w for (w, _s) in entries]
            stock = entries[0][1]
            srcr[plate][comp] = (wells, stock)
    return srcr


def _build_src_name_map(src: List[List[Any]], name_col : bool = False) -> Dict[str, Dict[str, str]]:
    """plate_barcode -> (well -> compound_name)"""
    out: Dict[str, Dict[str, str]] = {}
    for row in src[1:]:
        plate = as_str(row[5]).strip()
        well = as_str(row[0]).strip()
        if name_col:
            comp = as_str(row[3]).strip()
        else:
            comp = as_str(row[1]).strip()
            
        if plate == "" or well == "" or comp == "":
            continue
        out.setdefault(plate, {})[well] = comp
    return out

def _build_src_label_to_name_map(src: List[List[Any]]) -> Dict[str, Dict[str, str]]:
    """plate_barcode -> (well -> compound_name)"""
    out: Dict[str, Dict[str, str]] = {}
    for row in src[1:]:
        plate = as_str(row[5]).strip()
        label = as_str(row[1]).strip()
        comp = as_str(row[3]).strip()
            
        if plate == "" or label == "" or comp == "":
            continue
        out.setdefault(plate, {})[label] = comp
    return out


def _build_volumes(src: List[List[Any]]) -> Dict[str, float]:
    """key 'PLATE_WELL' -> volume_ul."""
    vols: Dict[str, float] = {}
    for row in src[1:]:
        plate = as_str(row[5]).strip()
        well = as_str(row[0]).strip()
        vol = as_number(row[4])
        if plate == "" or well == "" or vol is None:
            continue
        vols[f"{plate}_{well}"] = float(vol)
    return vols


def _dilute_volume_nl(
    well_volume_ul: float,
    src_stock: float,
    final_conc: float,
) -> float:
    desired_nl = 1000.0 * float(well_volume_ul) * float(final_conc) / float(src_stock)
    return round_to_increment(desired_nl, 2.5)


def _dilute_error_check(
    *,
    src_plate: str,
    src_name_map: Mapping[str, str],
    src_well_example: str,
    stock: float,
    final: float,
    well_volume_ul: float,
    rounded_nl: float,
    warnings_flag: List[bool],
) -> None:
    desired_nl = 1000.0 * float(well_volume_ul) * float(final) / float(stock)
    comp_name = src_name_map.get(src_well_example, src_well_example)

    # Warn on significant rounding error (matches WL assay script logic).
    if desired_nl != 0:
        rel_err = abs(desired_nl - float(rounded_nl)) / abs(desired_nl)
        if rel_err > 0.1:
            print(
                "Warning: Significant rounding error at low volume dispensing! "
                f"Compound: {comp_name} Stock: {stock} Final: {final} "
                f"Desired dispense volume: {desired_nl} Possible dispense volume: {rounded_nl}"
            )
            warnings_flag[0] = True

    # Abort on zero-volume transfers (Echo cannot execute 0 nL dispenses).
    if desired_nl != 0 and float(rounded_nl) == 0.0:
        raise PicklyPyConfigError(
            "Error: Zero-volume dispensing! "
            f"Compound: {comp_name} (source plate {src_plate}) Stock: {stock} Final: {final} "
            f"Desired dispense volume: {desired_nl} Possible dispense volume: {rounded_nl}"
        )



def _choose_source_well(
    *,
    src_plate: str,
    candidate_wells: Sequence[str],
    vol_nl: float,
    volumes_ul: Dict[str, float],
    minvol_ul: float,
    src_name_map: Mapping[str, str],
) -> str:
    need_ul = float(minvol_ul) + float(vol_nl) / 1000.0
    for w in candidate_wells:
        key = f"{src_plate}_{w}"
        have = float(volumes_ul.get(key, 0.0))
        if have >= need_ul:
            volumes_ul[key] = have - float(vol_nl) / 1000.0
            return w
    # Out of volume
    example = candidate_wells[0] if candidate_wells else ""
    comp_name = src_name_map.get(example, example)
    raise PicklyPyVolumeError(f"Error: Out of Volume for {comp_name}")


def _update_table_join(existing: List[List[str]], new: List[List[str]]) -> List[List[str]]:
    """Cell-wise join with ';' between non-empty existing/new, and newlines -> '+' in new."""
    out: List[List[str]] = []
    for r in range(PLATE_ROWS_384):
        row_out: List[str] = []
        for c in range(PLATE_COLS_384):
            a = existing[r][c]
            b = new[r][c].replace("\n", "+")
            if a != "" and b != "":
                row_out.append(a + ";" + b)
            else:
                row_out.append(a + b)
        out.append(row_out)
    return out


def _build_treatment_signatures_for_randomization(
    *,
    dst_sheet: List[List[Any]],
    additions: List[Addition],
    all_dst_barcodes: List[str],
    compoundlabels: List[Dict[str, str]],
) -> Dict[str, Dict[str, str]]:
    """Build per-plate per-well treatment signatures (dispenses only).

    This is used to compute replicate groups for well randomization.

    Notes
    -----
    - Signatures intentionally exclude the Layout map (if present) so that
      unused wells (no dispenses) remain unused after randomization.
    - In dose-response mode (Conc. map present) the per-well concentration
      is included in the signature so different doses are not considered
      replicates.
    """

    treatments_by_plate: Dict[str, Dict[str, str]] = {b: {} for b in all_dst_barcodes}

    for add in additions:
        plate_map_grid = _extract_plate_grid(dst_sheet, add.plate_map_row)
        if _is_grid_empty(plate_map_grid):
            continue

        dosing = add.conc_map_row is not None
        conc_map: Dict[str, float] = {}
        if dosing:
            conc_grid = _extract_plate_grid(dst_sheet, add.conc_map_row)
            for r in range(PLATE_ROWS_384):
                for c in range(PLATE_COLS_384):
                    num = as_number(conc_grid[r][c])
                    if num is None:
                        continue
                    if float(num) > 0:
                        conc_map[well_name(r + 1, c + 1)] = float(num)

        # Dose-response: label -> compound name (first occurrence).
        label_to_comp: Dict[str, str] = {}
        if dosing:
            for row in add.label_rows:
                lbl = as_str(row[0]).strip() if len(row) > 0 else ""
                comp = as_str(row[1]).strip() if len(row) > 1 else ""
                if lbl and comp and lbl not in label_to_comp:
                    label_to_comp[lbl] = comp

        dst_barcodes = all_dst_barcodes if add.barcodes_dst == ["*"] else add.barcodes_dst

        for r in range(PLATE_ROWS_384):
            for c in range(PLATE_COLS_384):
                cell = plate_map_grid[r][c]
                if cell.strip() == "":
                    continue
                w = well_name(r + 1, c + 1)

                if dosing and w not in conc_map:
                    # No dose specified -> no dispense.
                    continue

                labs = split_labels(cell)
                if not labs:
                    continue

                if dosing:
                    conc_final = conc_map[w]
                    conc_str = as_str(integer_chop(float(conc_final)))
                    parts = [f"{label_to_comp.get(lab, lab)}_{conc_str}" for lab in labs]
                    sig = "+".join(parts)
                else:
                    parts = [compoundlabels[add.index].get(lab, lab) for lab in labs]
                    sig = "+".join(parts)

                if sig.strip() == "":
                    continue

                for bc in dst_barcodes:
                    prev = treatments_by_plate[bc].get(w, "")
                    treatments_by_plate[bc][w] = prev + (";" if prev else "") + sig

    return treatments_by_plate


def generate_assay_picklists(
    xlsx_path: str | Path,
    *,
    pause_on_exit: bool = False,
    randomize_wells: bool = False,
    name_column: bool = False,
) -> None:
    """Main API: generate assay picklists and Plate:Works sidecar files."""
    _print_banner()

    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        raise PicklyPyConfigError("Picklist file not found or cannot be imported!")

    wb = load_workbook(xlsx_path, data_only=True)

    # WL uses first two worksheets; we try names first.
    ws_src = wb[wb.sheetnames[0]]
    ws_dst = wb[wb.sheetnames[1]] if len(wb.sheetnames) > 1 else None
    if ws_dst is None:
        raise PicklyPyConfigError(f"Worksheet not found: {xlsx_path}")

    src = sheet_to_matrix(ws_src)
    dst = sheet_to_matrix(ws_dst)

    if not src or not dst:
        raise PicklyPyConfigError(f"Worksheet not found: {xlsx_path}")

    # Validate SRC.
    _validate_src_sheet(src)

    print(f"{xlsx_path} loaded successfully.")

    header = _parse_header(dst)

    # Min volume in SRC! G1 (row 1 col 7)
    minvol = as_number(src[0][6] if len(src[0]) > 6 else None)
    if minvol is None or not math.isfinite(float(minvol)):
        raise PicklyPyConfigError(
            "Error: missing minimal Echo Source well volume in 'src' worksheet G1!"
        )
    minvol_ul = float(minvol)

    additions, all_dst_barcodes = _parse_additions(dst, header.well_volume_ul_default)

    # Survey volumes.
    _apply_survey_to_src(src, xlsx_path.parent, used_source_barcodes=[a.barcode_src for a in additions])

    # Build source mapping / name map / volumes.
    srcr = _build_srcr(src)

    # Force library-style naming if PickliPy.Screen header is present
    if src[0][3] == "Name in Library":
        name_column = True

    # Ensure every source barcode referenced exists.
    for a in additions:
        if a.barcode_src not in srcr:
            raise PicklyPyConfigError(
                f"The {a.barcode_src} barcode does not exist in the SRC worksheet Col F!"
            )

    src_name = _build_src_name_map(src,name_col = name_column)
    if name_column:
        src_label_to_name = _build_src_label_to_name_map(src)
    else:
        src_label_to_name = ()
    
    volumes_ul = _build_volumes(src)

    # Validate label list compounds exist per addition.
    for add in additions:
        missing: List[str] = []
        plate_map = srcr.get(add.barcode_src, {})
        for row in add.label_rows:
            comp = as_str(row[1]).strip()
            if comp == "":
                continue
            if comp not in plate_map:
                missing.append(comp)
        if missing:
            missing_u = ", ".join(sorted(set(missing)))
            raise PicklyPyConfigError(f"Missing definitons on SRC: {missing_u}")

    # Build compoundlabels mapping used for conditions/merge files.
    compoundlabels: List[Dict[str, str]] = []
    for add in additions:
        label_to_parts: Dict[str, List[str]] = {}
        for row in add.label_rows:
            lbl = as_str(row[0]).strip()
            comp = as_str(row[1]).strip() 
            
            if name_column:
                comp = src_label_to_name.get(add.barcode_src).get(comp,comp)

                               
            conc = row[2] if len(row) > 2 else ""
            if lbl == "" or comp == "":
                continue
            conc_num = as_number(conc)
            conc_str = as_str(integer_chop(conc_num) if conc_num is not None else conc)
            label_to_parts.setdefault(lbl, []).append(f"{comp}_{conc_str}")
        label_to_joined = {k: "+".join(v) for k, v in label_to_parts.items()}
        compoundlabels.append(label_to_joined)

    # Layout map (optional).
    layout_rows = _find_rows_with_label(dst, "Layout:")
    layout_grid: Optional[List[List[str]]] = None
    base_conditions = [["" for _ in range(PLATE_COLS_384)] for _ in range(PLATE_ROWS_384)]
    if layout_rows:
        layout_grid = _extract_plate_grid(dst, layout_rows[0])
        base_conditions = _update_table_join(base_conditions, layout_grid)

    # Layout label by well (used for group-wise randomization).
    layout_by_well: Dict[str, str] = {}
    if layout_grid is not None:
        for r in range(PLATE_ROWS_384):
            for c in range(PLATE_COLS_384):
                layout_by_well[well_name(r + 1, c + 1)] = layout_grid[r][c].strip()

    # Optional: destination-well randomization (per destination barcode).
    randomization_maps: Dict[str, Dict[str, str]] = {}
    if randomize_wells:
        print("****** Well randomization enabled (per destination plate barcode) - this may take a minute - random seed is based on design file name ******")
        treatments_by_plate = _build_treatment_signatures_for_randomization(
            dst_sheet=dst,
            additions=additions,
            all_dst_barcodes=all_dst_barcodes,
            compoundlabels=compoundlabels,
        )

        for bc in all_dst_barcodes:
            seed = seed_from_text(f"{xlsx_path.name}|{bc}")
            try:
                randomization_maps[bc] = optimize_plate_well_map(
                    treatment_by_well=treatments_by_plate.get(bc, {}),
                    layout_by_well=layout_by_well if layout_by_well else None,
                    seed=seed,
                    attempts=100,
                    forbid_adjacent_replicates=True,
                )
            except ValueError as e:
                raise PicklyPyConfigError(
                    f"Well randomization failed for destination plate {bc}: {e}"
                ) from e

        # Brief summary.
        for bc in all_dst_barcodes:
            n_used = len(treatments_by_plate.get(bc, {}))
            n_mapped = len(randomization_maps.get(bc, {}))
            print(f"Randomization map for {bc}: {n_mapped} wells mapped ({n_used} used wells).")

    # Per destination plate accumulated conditions.
    allplates_labels: Dict[str, List[List[str]]] = {b: copy.deepcopy(base_conditions) for b in all_dst_barcodes}
    allplates_compounds: Dict[str, List[List[str]]] = {b: copy.deepcopy(base_conditions) for b in all_dst_barcodes}

    # State for merging consecutive additions with the same picklist name.
    previous_picklist_name = ""
    previous_csv_rows: List[List[Any]] = []

    totaluse: List[Tuple[str, str, float]] = []  # (compound_name, plate_well_key, used_ul)
    warnings = [False]

    for add in additions:
        print(f"****** {add.picklist_name}.csv ******")
        well_volume_ul = add.well_volume_ul
        print(f"Well volume = {well_volume_ul} ul.")

        plate_map_grid = _extract_plate_grid(dst, add.plate_map_row)
        if _is_grid_empty(plate_map_grid):
            print("Skipping empty Plate Map...")
            out_path = xlsx_path.parent / f"{add.picklist_name}.csv"
            write_text(out_path, format_csv_rows(previous_csv_rows, quote=False))
            print(f"Saving: {out_path}")
            continue

        # Build list of (dstwell, label) from plate map.
        c_pairs: List[Tuple[str, str]] = []
        for r in range(PLATE_ROWS_384):
            for c in range(PLATE_COLS_384):
                cell = plate_map_grid[r][c]
                if cell.strip() == "":
                    continue
                dstwell = well_name(r+1,c+1)
                for lab in split_labels(cell):
                    c_pairs.append((dstwell, lab))

        # All labels used must be defined in label list.
        defined_labels = {as_str(row[0]).strip() for row in add.label_rows if as_str(row[0]).strip() != ""}
        used_labels = {lab for (_w, lab) in c_pairs}
        missing_labels = sorted(used_labels - defined_labels)
        if missing_labels:
            raise PicklyPyConfigError(f"Error: Missing label definition for: {missing_labels}")

        # Dose-response mode?
        dosing = add.conc_map_row is not None
        conc_map: Dict[str, float] = {}
        if dosing:
            conc_grid = _extract_plate_grid(dst, add.conc_map_row)
            for r in range(PLATE_ROWS_384):
                for c in range(PLATE_COLS_384):
                    v = conc_grid[r][c]
                    num = as_number(v)
                    if num is None:
                        continue
                    if float(num) > 0:
                        dstwell = well_name(r+1,c+1)
                        conc_map[dstwell] = float(num)

        # Prepare label list decoding.
        # Map plate-map label -> list of (src_wells_list, stock_conc, final_conc).
        label_defs: Dict[str, List[Tuple[List[str], float, float]]] = {}
        for row in add.label_rows:
            lbl = as_str(row[0]).strip()
            comp = as_str(row[1]).strip()
            conc = as_number(row[2] if len(row) > 2 else None)
            if lbl == "" or comp == "":
                continue
            if dosing:
                # Concentration in label list is ignored for dispensing.
                pass
            if conc is None:
                if not dosing:
                    raise PicklyPyConfigError(
                        f"Error: missing (or non-numeric) final concentration for label '{lbl}' "
                        f"in the DST label list for picklist '{add.picklist_name}'."
                    )
                conc = 0.0
            srcinfo = srcr[add.barcode_src][comp]
            label_defs.setdefault(lbl, []).append((srcinfo[0], srcinfo[1], float(conc)))

        # In dosing mode, mimic WL behavior: only the first mapping per label is used.
        if dosing:
            label_defs = {k: [v[0]] for k, v in label_defs.items() if v}

        # Determine destination barcodes for this addition.
        dst_barcodes = all_dst_barcodes if add.barcodes_dst == ["*"] else add.barcodes_dst

        # Start CSV accumulator for this file.
        if previous_picklist_name == add.picklist_name:
            csv_accum: List[List[Any]] = list(previous_csv_rows)
            print("Merging picklist with previous.")
        else:
            csv_accum = []

        # Barcode-independent plate map of compound labels for this addition (used for metadata).
        # If destination well randomization is enabled, this grid will be permuted per destination barcode.
        platemap_with_compounds_base: List[List[str]] = []
        for r in range(PLATE_ROWS_384):
            row_out: List[str] = []
            for c in range(PLATE_COLS_384):
                cell = plate_map_grid[r][c]
                if cell.strip() == "":
                    row_out.append("")
                    continue
                parts: List[str] = []
                for lab in split_labels(cell):
                    parts.append(compoundlabels[add.index].get(lab, lab))
                row_out.append("+".join(parts))
            platemap_with_compounds_base.append(row_out)

        # For each destination plate barcode, generate dispenses.
        for dst_barcode in dst_barcodes:
            print(f"Dispensing from: {add.barcode_src}")
            print(f"Dispensing into: {dst_barcode}")

            # Destination-well randomization (per destination barcode): map original well -> randomized well.
            well_map = randomization_maps.get(dst_barcode, {}) if randomize_wells else {}

            # Update conditions maps for this destination plate (apply randomization to the plate maps).
            plate_map_grid_use = apply_well_map_to_str_grid(plate_map_grid, well_map) if well_map else plate_map_grid
            platemap_with_compounds_use = (
                apply_well_map_to_str_grid(platemap_with_compounds_base, well_map)
                if well_map
                else platemap_with_compounds_base
            )

            allplates_labels[dst_barcode] = _update_table_join(allplates_labels[dst_barcode], plate_map_grid_use)
            allplates_compounds[dst_barcode] = _update_table_join(
                allplates_compounds[dst_barcode], platemap_with_compounds_use
            )

            # Also apply randomization to dispensing targets (before the greedy travel optimization).
            c_pairs_use = (
                [(well_map.get(dstwell, dstwell), lab) for (dstwell, lab) in c_pairs]
                if well_map
                else c_pairs
            )
            conc_map_use = apply_well_map_to_well_dict(conc_map, well_map) if (dosing and well_map) else conc_map

            # Expand dispensing events.
            events: List[Tuple[str, List[str], float]] = []  # (dstwell, candidate_src_wells, vol_nl)
            if dosing:
                for dstwell, lab in c_pairs_use:
                    if dstwell not in conc_map_use:
                        continue
                    conc_final = conc_map_use[dstwell]
                    if lab not in label_defs:
                        continue
                    src_wells, stock, _ignored = label_defs[lab][0]
                    vol_nl = _dilute_volume_nl(well_volume_ul, stock, conc_final)
                    _dilute_error_check(
                        src_plate=add.barcode_src,
                        src_name_map=src_name[add.barcode_src],
                        src_well_example=src_wells[0],
                        stock=stock,
                        final=conc_final,
                        well_volume_ul=well_volume_ul,
                        rounded_nl=vol_nl,
                        warnings_flag=warnings,
                    )
                    events.append((dstwell, src_wells, vol_nl))
            else:
                for dstwell, lab in c_pairs_use:
                    for (src_wells, stock, final_conc) in label_defs.get(lab, []):
                        if final_conc == 0:
                            # A 0 concentration entry would produce 0 volume; WL would error.
                            pass
                        vol_nl = _dilute_volume_nl(well_volume_ul, stock, final_conc)
                        _dilute_error_check(
                            src_plate=add.barcode_src,
                            src_name_map=src_name[add.barcode_src],
                            src_well_example=src_wells[0],
                            stock=stock,
                            final=final_conc,
                            well_volume_ul=well_volume_ul,
                            rounded_nl=vol_nl,
                            warnings_flag=warnings,
                        )
                        events.append((dstwell, src_wells, vol_nl))

            # Group by candidate source wells list (compound).
            grouped: Dict[Tuple[str, ...], List[Tuple[str, List[str], float]]] = {}
            order_keys: List[Tuple[str, ...]] = []
            for ev in events:
                key = tuple(ev[1])
                if key not in grouped:
                    grouped[key] = []
                    order_keys.append(key)
                grouped[key].append(ev)

            groups = [grouped[k] for k in order_keys]

            # Reorder groups by source travel.
            groups = reorder_source_groups(groups)

            # Reorder destinations within each group.
            last_pos = (1, 1)
            ordered_events: List[Tuple[str, List[str], float]] = []
            for g in groups:
                g2, last_pos = reorder_destinations_within_group(g, last_pos)
                ordered_events.extend(g2)

            # Choose concrete source wells and group by chosen well.
            chosen_events: List[Tuple[str, str, float]] = []
            for dstwell, candidate_wells, vol_nl in ordered_events:
                chosen = _choose_source_well(
                    src_plate=add.barcode_src,
                    candidate_wells=candidate_wells,
                    vol_nl=vol_nl,
                    volumes_ul=volumes_ul,
                    minvol_ul=minvol_ul,
                    src_name_map=src_name[add.barcode_src],
                )
                chosen_events.append((dstwell, chosen, vol_nl))

            by_srcwell: Dict[str, List[Tuple[str, str, float]]] = {}
            by_srcwell_order: List[str] = []
            for ev in chosen_events:
                w = ev[1]
                if w not in by_srcwell:
                    by_srcwell[w] = []
                    by_srcwell_order.append(w)
                by_srcwell[w].append(ev)

            # Print volumes used per well, and build CSV rows.
            for src_well in by_srcwell_order:
                evs = by_srcwell[src_well]
                used_ul = sum(v for (_dw, _sw, v) in evs) / 1000.0
                comp_name = src_name[add.barcode_src].get(src_well, src_well)
                key = f"{add.barcode_src}_{src_well}"
                totaluse.append((comp_name, key, float(used_ul)))

            print("Volumes used:")
            for src_well in by_srcwell_order:
                comp_name = src_name[add.barcode_src].get(src_well, src_well)
                key = f"{add.barcode_src}_{src_well}"
                used_ul = sum(v for (_dw, sw, v) in by_srcwell[src_well]) / 1000.0
                print(f"{comp_name}   {key}   {used_ul}")

            for src_well in by_srcwell_order:
                evs = by_srcwell[src_well]
                row: List[Any] = [add.barcode_src, src_well, dst_barcode]
                for dstwell, _sw, vol_nl in evs:
                    row.extend([vol_nl, dstwell])
                csv_accum.append(row)

        # Persist for next addition.
        previous_picklist_name = add.picklist_name
        previous_csv_rows = csv_accum

        out_path = xlsx_path.parent / f"{add.picklist_name}.csv"
        write_text(out_path, format_csv_rows(csv_accum, quote=False))
        print(f"Saving: {out_path}")

    # Update SRC sheet volumes and save picklist_after_dispense.xlsx
    for i in range(1, len(src)):
        row = src[i]
        plate = as_str(row[5]).strip()
        well = as_str(row[0]).strip()
        if plate == "" or well == "":
            continue
        key = f"{plate}_{well}"
        if key in volumes_ul:
            row[4] = float(volumes_ul[key])

    out_xlsx = xlsx_path.parent / "picklist_after_dispense.xlsx"
    _save_updated_workbook(xlsx_path, out_xlsx, src_sheet_name=ws_src.title, src_matrix=src)

    # Inventory and process files
    print("***** Plateworks Tables ******")

    inv_src_rows = [make_inventory_row(b, header.rack_src, idx + 1) for idx, b in enumerate(sorted(set(a.barcode_src for a in additions)))]
    inv_dst_rows = [make_inventory_row(b, header.rack_dst, idx + 1) for idx, b in enumerate(all_dst_barcodes)]

    if header.inventory_src == header.inventory_dst:
        inv_rows = inv_src_rows + inv_dst_rows
        inv_path = xlsx_path.parent / f"{header.inventory_src}.csv"
        write_text(inv_path, format_csv_rows(inv_rows, quote=False))
        print(f"***** Inventory = {inv_path.name}")
    else:
        inv_path = xlsx_path.parent / f"{header.inventory_src}.csv"
        write_text(inv_path, format_csv_rows(inv_src_rows, quote=False))
        print(f"***** Inventory = {inv_path.name}")

        inv_path2 = xlsx_path.parent / f"{header.inventory_dst}.csv"
        write_text(inv_path2, format_csv_rows(inv_dst_rows, quote=False))
        print(f"***** Inventory = {inv_path2.name}")

    proc_src_path = xlsx_path.parent / f"{header.process_src}.txt"
    proc_dst_path = xlsx_path.parent / f"{header.process_dst}.txt"

    proc_src_barcodes = sorted(set(a.barcode_src for a in additions))
    proc_dst_barcodes = all_dst_barcodes

    write_text(proc_src_path, "\n".join(proc_src_barcodes))
    write_text(proc_dst_path, "\n".join(proc_dst_barcodes))

    print(f"***** Process_SRC = {proc_src_path.name}")
    print(f"***** Process_DST = {proc_dst_path.name}")

    # Per-plate conditions list files.
    for plate in all_dst_barcodes:
        labels_flat = [x for row in allplates_labels[plate] for x in row if x != ""]
        comps_flat = [x for row in allplates_compounds[plate] for x in row if x != ""]

        out_labels = xlsx_path.with_name(f"{xlsx_path.name}_{plate}_conditionslist labels.txt")
        out_comps = xlsx_path.with_name(f"{xlsx_path.name}_{plate}_conditionslist compounds.txt")

        write_text(out_labels, "\n".join(labels_flat))
        write_text(out_comps, "\n".join(comps_flat))

    # Plate-map exports (grid format), similar to PicklyPy.Screen.
    def _write_plate_maps(path: Path, tables: Dict[str, List[List[str]]]) -> None:
        lines: List[str] = []
        for plate in all_dst_barcodes:
            lines.append(plate)
            tbl = tables[plate]
            for r in range(PLATE_ROWS_384):
                lines.append("\t".join(tbl[r][c] for c in range(PLATE_COLS_384)))
        write_text(path, "\n".join(lines))

    platemap_labels_path = xlsx_path.with_name(f"{xlsx_path.name}_plate map labels.txt")
    platemap_compounds_path = xlsx_path.with_name(f"{xlsx_path.name}_plate map compounds.txt")
    _write_plate_maps(platemap_labels_path, allplates_labels)
    _write_plate_maps(platemap_compounds_path, allplates_compounds)

    # Merge tables
    def _merge_rows(table: List[List[str]], plate: str) -> List[List[str]]:
        rows: List[List[str]] = []
        for r in range(PLATE_ROWS_384):
            for c in range(PLATE_COLS_384):
                v = table[r][c]
                if v == "":
                    continue
                rows.append([well_name(r+1,c+1), plate, v])
        return rows

    merge_labels_rows: List[List[str]] = []
    merge_comp_rows: List[List[str]] = []
    for plate in all_dst_barcodes:
        merge_labels_rows.extend(_merge_rows(allplates_labels[plate], plate))
        merge_comp_rows.extend(_merge_rows(allplates_compounds[plate], plate))

    merge_labels_path = xlsx_path.with_name(f"{xlsx_path.name}_table merge labels.txt")
    merge_comp_path = xlsx_path.with_name(f"{xlsx_path.name}_table merge compounds.txt")

    write_text(merge_labels_path, format_csv_rows(merge_labels_rows, quote=True))
    write_text(merge_comp_path, format_csv_rows(merge_comp_rows, quote=True))

    # Summary
    print("****** Summary: Total Usage ******")
    # Group by plate_well key
    usage_by_key: Dict[str, List[Tuple[str, str, float]]] = {}
    for comp, key, used in totaluse:
        usage_by_key.setdefault(key, []).append((comp, key, used))

    for key, items in usage_by_key.items():
        comp_name = items[0][0]
        used_total = sum(x[2] for x in items)
        remains = volumes_ul.get(key, 0.0)
        print(f"{comp_name}\t\t{key}\t{used_total:.2f}\tremains (ul): {remains:.2f}")

    if warnings[0]:
        print("Finished with warnings. Revise source file if needed!")
    else:
        print("Picklists passed all checks. Ready.")


def _save_updated_workbook(
    original_path: Path,
    out_path: Path,
    *,
    src_sheet_name: str,
    src_matrix: List[List[Any]],
) -> None:
    """Save a copy of the original workbook with updated SRC volume column (E)."""
    wb = openpyxl.load_workbook(original_path, data_only=False)
    if src_sheet_name not in wb.sheetnames:
        # Fallback: first sheet
        ws = wb.worksheets[0]
    else:
        ws = wb[src_sheet_name]

    # Update column E (5) for rows 2..N according to src_matrix.
    for r in range(2, len(src_matrix) + 1):
        # src_matrix is 0-based.
        row = src_matrix[r - 1]
        if len(row) < 5:
            continue
        ws.cell(row=r, column=5).value = row[4]

    wb.save(out_path)


def main(argv: Optional[Sequence[str]] = None) -> int:
    import argparse
    
    def _str2bool(v: object) -> bool:
        # Accept common boolean spellings; require an explicit value.
        if isinstance(v, bool):
            return v
        s = str(v).strip().lower()
        if s in {"true", "t", "1", "yes", "y"}:
            return True
        if s in {"false", "f", "0", "no", "n"}:
            return False
        raise argparse.ArgumentTypeError(
            "Expected a boolean value for --randomize-wells (true/false)."
        )

    parser = argparse.ArgumentParser(prog="PicklyPy.Assay", description="Generate assay picklists from an Excel design file.")
    parser.add_argument("-f", "--file", dest="xlsx", required=True, help="Path to the design .xlsx file")
    parser.add_argument("--no-pause", action="store_true", help="Do not wait for Enter at the end")
    parser.add_argument(
        "-r","--randomize_wells",
        dest="randomize_wells",
        type=_str2bool,
        default=False,
        metavar="{true|false}",
        help=(
            "Randomize destination well positions (treatments are shuffled, the set of used wells is unchanged). "
            "Randomization is applied independently for each destination plate barcode. If a Layout: map is present "
            "in the DST worksheet, shuffling is performed independently within each layout label group."
        )
    )
    parser.add_argument(
        "-n","--name_column",
        dest="name_column",
        type=_str2bool,
        default=False,
        metavar="{true|false}",
        help=(
            "Use compound names from column D (similary to PickliPy.Screen in metadata files."
            "This allows using labels/slots instead of compound names in the label list of the DST sheet, and assign Compound names to these labels/slots in the SRC sheet."

        )
    )

    args = parser.parse_args(list(argv) if argv is not None else None)

    # Mimic the original Wolfram script's behavior: pause at the end unless the
    # user explicitly opts out.
    return run_with_user_facing_errors(
        lambda: generate_assay_picklists(args.xlsx, randomize_wells=args.randomize_wells, name_column=args.name_column),
        pause_on_exit=not args.no_pause,
    )



if __name__ == "__main__":
    raise SystemExit(main())
