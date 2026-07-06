from __future__ import annotations

import math
import os
import re
import sys
import traceback
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, Iterator, List, Mapping, MutableMapping, Optional, Sequence, Tuple

import openpyxl


# -----------------------------
# Exceptions
# -----------------------------


class PicklyPyError(RuntimeError):
    """Base class for user-facing errors."""


class PicklyPyConfigError(PicklyPyError):
    """Raised when the input design file is missing required fields or is malformed."""


class PicklyPyVolumeError(PicklyPyError):
    """Raised when an operation would deplete a source well below the minimum volume."""


# -----------------------------
# Constants
# -----------------------------


PLATE_ROWS_384 = 16
PLATE_COLS_384 = 24
PLATE_ROWS_1536 = 32
PLATE_COLS_1536 = 48


# -----------------------------
# Utility functions
# -----------------------------


def _is_blank(v: Any) -> bool:
    if v is None:
        return True
    if isinstance(v, str):
        return v.strip() == ""
    return False


def as_str(v: Any) -> str:
    """Convert Excel cell values into the string behavior expected by the WL scripts."""
    if v is None:
        return ""
    if isinstance(v, str):
        return v
    # Keep integers without trailing .0
    if isinstance(v, (int,)):
        return str(v)
    if isinstance(v, float):
        if math.isfinite(v) and abs(v - round(v)) < 1e-9:
            return str(int(round(v)))
        return str(v)
    return str(v)


def as_number(v: Any) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip()
        if s == "":
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def integer_chop(n: Any) -> Any:
    """If n is a float extremely close to an integer, convert it to an int."""
    if not isinstance(n, (int, float)):
        return n
    if isinstance(n, bool):
        return n
    r = round(float(n))
    if math.isfinite(float(n)) and abs(float(n) - r) < 1e-9:
        return int(r)
    return n


def round_half_away_from_zero(x: float) -> int:
    """Round to nearest integer with ties away from zero (Mathematica-style)."""
    if x >= 0:
        return int(math.floor(x + 0.5))
    return int(math.ceil(x - 0.5))


def round_to_increment(x: float, inc: float) -> float:
    """Round x to the nearest multiple of inc, with tie-breaking away from zero."""
    if inc <= 0:
        raise ValueError("inc must be > 0")
    return round_half_away_from_zero(x / inc) * inc


_LABEL_SPLIT_RE = re.compile(r"[\+\,;\n\r]+")


def split_labels(cell_text: str) -> List[str]:
    """Split a plate-map cell into label tokens.

    Matches the WL scripts: separators are + , ; and newlines.
    """
    if cell_text is None:
        return []
    s = as_str(cell_text)
    if s.strip() == "":
        return []
    parts = [p.strip() for p in _LABEL_SPLIT_RE.split(s) if p.strip() != ""]
    return parts


def excel_col_to_number(col: str) -> int:
    """Convert Excel column letters (A, Z, AA, ...) to 1-based index."""
    s = col.strip().upper()
    if not s or not re.fullmatch(r"[A-Z]+", s):
        raise ValueError(f"Invalid Excel column: {col!r}")
    n = 0
    for ch in s:
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n


def number_to_excel_col(n: int) -> str:
    """Convert 1-based index to Excel column letters."""
    if n <= 0:
        raise ValueError("n must be >= 1")
    out = []
    x = n
    while x > 0:
        x, rem = divmod(x - 1, 26)
        out.append(chr(ord("A") + rem))
    return "".join(reversed(out))


_WELL_RE = re.compile(r"^(?P<row>[A-Za-z]+)(?P<col>\d+)$")


def parse_well_name(well: str) -> Tuple[int, int]:
    """Parse a well name like A1 or AA12 into (row_index, col_index), 1-based."""
    m = _WELL_RE.match(well.strip())
    if not m:
        raise ValueError(f"Invalid well name: {well!r}")
    row_letters = m.group("row").upper()
    col = int(m.group("col"))
    row = excel_col_to_number(row_letters)  # supports multi-letter rows
    return row, col


def well_name(row: int, col: int) -> str:
    """Convert (row_index, col_index) to a well name.

    For 384-well plates, rows are A-P; we still support multi-letter rows.
    """
    return f"{number_to_excel_col(row)}{col}"


def euclidean(a: Tuple[int, int], b: Tuple[int, int]) -> float:
    return math.hypot(a[0] - b[0], a[1] - b[1])


def _nearest_neighbor_path(
    points: List[Tuple[int, int]],
    start_idx: int,
    end_idx: int,
) -> List[int]:
    """Greedy path visiting all points once, with fixed start and end.

    This is intended to emulate Mathematica FindShortestTour[..., Method->"Greedy"]
    used with specified first/last indices.

    The algorithm:
      - Start at start_idx.
      - Repeatedly go to nearest unvisited point, but keep end_idx reserved
        for the final step.
      - Finish at end_idx.

    Returns the ordered list of indices into `points`.
    """
    if not points:
        return []
    if len(points) == 1:
        return [0]

    n = len(points)
    if not (0 <= start_idx < n and 0 <= end_idx < n):
        raise ValueError("start_idx/end_idx out of range")

    unvisited = set(range(n))
    path = [start_idx]
    unvisited.remove(start_idx)

    # Reserve the end node if it's not the start node.
    reserved_end = end_idx if end_idx in unvisited else None

    cur = start_idx
    while unvisited:
        # If the only remaining node is the reserved end, finish.
        if reserved_end is not None and unvisited == {reserved_end}:
            path.append(reserved_end)
            unvisited.remove(reserved_end)
            break

        # Candidate pool excludes reserved end (if reserved).
        candidates = list(unvisited)
        if reserved_end is not None and reserved_end in unvisited:
            candidates = [i for i in candidates if i != reserved_end]

        if not candidates:
            # Only reserved end remains.
            if reserved_end is not None:
                path.append(reserved_end)
                unvisited.remove(reserved_end)
            break

        # Choose nearest neighbor.
        cur_pt = points[cur]
        nxt = min(candidates, key=lambda i: euclidean(cur_pt, points[i]))
        path.append(nxt)
        unvisited.remove(nxt)
        cur = nxt

    return path


def reorder_source_groups(
    groups: List[List[Tuple[str, List[str], float]]],
    origin: Tuple[int, int] = (1, 1),
) -> List[List[Tuple[str, List[str], float]]]:
    """Reorder groups (each group is events for one compound) by source-well travel."""
    if not groups:
        return groups

    # Use first source well of each group as representative.
    reps: List[Tuple[int, int]] = []
    for g in groups:
        if not g:
            reps.append(origin)
            continue
        first_well = g[0][1][0]  # first event, list of wells -> first well
        reps.append(parse_well_name(first_well))

    # Determine start/end similar to WL: nearest and farthest from origin.
    dists = [euclidean(origin, p) for p in reps]
    start_idx = int(min(range(len(groups)), key=lambda i: dists[i]))
    end_idx = int(max(range(len(groups)), key=lambda i: dists[i]))

    order = _nearest_neighbor_path(reps, start_idx=start_idx, end_idx=end_idx)
    return [groups[i] for i in order]


def reorder_destinations_within_group(
    events: List[Tuple[str, List[str], float]],
    last_pos: Tuple[int, int],
) -> Tuple[List[Tuple[str, List[str], float]], Tuple[int, int]]:
    """Reorder events within a group by destination-well travel, tracking last_pos."""
    if not events:
        return events, last_pos

    pts = [parse_well_name(ev[0]) for ev in events]  # destination wells
    dists = [euclidean(last_pos, p) for p in pts]
    start_idx = int(min(range(len(events)), key=lambda i: dists[i]))
    end_idx = int(max(range(len(events)), key=lambda i: dists[i]))

    order = _nearest_neighbor_path(pts, start_idx=start_idx, end_idx=end_idx)
    reordered = [events[i] for i in order]

    # Update last_pos to the last destination of this group, like WL.
    new_last = parse_well_name(reordered[-1][0])
    return reordered, new_last


def quote_csv_field(x: Any) -> str:
    """Quote a CSV field the same way the WL scripts do for merge tables."""
    s = as_str(x)
    if "," in s or '"' in s:
        return '"' + s.replace('"', '""') + '"'
    return s


def write_text(path: Path, text: str) -> None:
    path.write_text(text, encoding="utf-8", newline="\n")


def tally_ordered(items: Sequence[str]) -> List[Tuple[str, int]]:
    """Mathematica-like Tally that preserves first-seen order."""
    counts: Dict[str, int] = {}
    order: List[str] = []
    for it in items:
        if it not in counts:
            counts[it] = 1
            order.append(it)
        else:
            counts[it] += 1
    return [(k, counts[k]) for k in order]


# -----------------------------
# Excel helpers
# -----------------------------


def load_workbook(path: Path, data_only: bool = True) -> openpyxl.Workbook:
    return openpyxl.load_workbook(path, data_only=data_only)


def sheet_to_matrix(ws: openpyxl.worksheet.worksheet.Worksheet, min_cols: int = 25) -> List[List[Any]]:
    """Read a worksheet into a rectangular list-of-lists.

    - Empty cells become "" (empty string) to match the WL scripts.
    - Ensures at least `min_cols` columns exist (plate maps use up to 25 columns).

    The WL Import tends to trim trailing empty rows/cols; here we approximate that
    by finding the last non-empty row/col.
    """

    max_row = ws.max_row or 1
    max_col = max(ws.max_column or 1, min_cols)

    # Determine the last row that has any non-empty value in the first max_col columns.
    last_row = 0
    for r in range(1, max_row + 1):
        row_has = False
        for c in range(1, max_col + 1):
            v = ws.cell(r, c).value
            if not _is_blank(v):
                row_has = True
                break
        if row_has:
            last_row = r

    if last_row == 0:
        last_row = 1

    # Determine last col with any non-empty value within last_row rows.
    last_col = 0
    for c in range(1, max_col + 1):
        col_has = False
        for r in range(1, last_row + 1):
            v = ws.cell(r, c).value
            if not _is_blank(v):
                col_has = True
                break
        if col_has:
            last_col = c

    last_col = max(last_col, min_cols)

    matrix: List[List[Any]] = []
    for r in range(1, last_row + 1):
        row: List[Any] = []
        for c in range(1, last_col + 1):
            v = ws.cell(r, c).value
            if v is None:
                v = ""
            row.append(v)
        matrix.append(row)

    return matrix


def get_sheet_matrix_by_name_or_index(
    wb: openpyxl.Workbook,
    preferred_names: Sequence[str],
    index_fallback: int,
) -> Tuple[str, List[List[Any]]]:
    """Return (sheet_name, matrix), trying names first then index."""
    for name in preferred_names:
        if name in wb.sheetnames:
            ws = wb[name]
            return ws.title, sheet_to_matrix(ws)
    # fallback by index
    if index_fallback < 0 or index_fallback >= len(wb.worksheets):
        raise PicklyPyConfigError(
            f"Worksheet not found. Tried names={preferred_names} and index={index_fallback}."
        )
    ws = wb.worksheets[index_fallback]
    return ws.title, sheet_to_matrix(ws)


def normalize_row_length(rows: List[List[Any]], min_cols: int) -> List[List[Any]]:
    """Pad rows with empty strings so all rows have at least min_cols columns."""
    out: List[List[Any]] = []
    for r in rows:
        if len(r) < min_cols:
            out.append(r + [""] * (min_cols - len(r)))
        else:
            out.append(r)
    return out


# -----------------------------
# Survey file parsing
# -----------------------------


@dataclass(frozen=True)
class SurveyVolumes:
    """Parsed Echo plate survey volumes.

    Keys are (barcode, well_name) -> volume_ul.
    """

    volumes_ul: Mapping[Tuple[str, str], float]
    unknown_barcodes: Tuple[Path, ...] = ()

    def get(self, barcode: str, well: str) -> Any:
        return self.volumes_ul.get((barcode, well), None)


def parse_plate_survey_xml(path: Path) -> Tuple[Optional[str], Dict[Tuple[str, str], float]]:
    """Parse one *_platesurvey.xml file.

    Returns (barcode, volumes) where volumes is mapping (barcode, well) -> volume_ul.
    If barcode cannot be read, returns (None, {}).
    """
    try:
        tree = ET.parse(path)
    except Exception as e:
        raise PicklyPyError(f"Failed to parse survey XML: {path}: {e}")

    root = tree.getroot()

    # Find the first platesurvey element and its barcode attribute.
    barcode = None
    for el in root.iter():
        if el.tag.lower().endswith("platesurvey"):
            barcode = el.attrib.get("barcode")
            break

    # Collect well volumes.
    vols: Dict[Tuple[str, str], float] = {}
    if barcode is None:
        return None, vols

    # Typical structure: <w n="A1" vl="12.3" ... />
    for w in root.iter():
        if w.tag.lower().endswith("w"):
            well = w.attrib.get("n")
            vl = w.attrib.get("vl")
            if well is None or vl is None:
                continue
            try:
                vol = float(vl)
            except ValueError:
                continue
            vols[(barcode, well)] = vol

    return barcode, vols


def load_survey_volumes(folder: Path) -> SurveyVolumes:
    """Load all *_platesurvey.xml files in a folder and merge volumes."""
    merged: Dict[Tuple[str, str], float] = {}
    unknown: List[Path] = []

    for p in sorted(folder.glob("*_platesurvey.xml")):
        barcode, vols = parse_plate_survey_xml(p)
        if barcode is None:
            unknown.append(p)
            continue
        # Later files overwrite earlier ones if the same (barcode,well) appears.
        merged.update(vols)

    return SurveyVolumes(volumes_ul=merged, unknown_barcodes=tuple(unknown))


# -----------------------------
# Plate:Works helpers
# -----------------------------


def make_inventory_row(barcode: str, rack: int, pos: int) -> List[Any]:
    """Create one row for Plate:Works inventory CSV (Echo templates)."""
    # Mirrors the WL: {barcode,"","",1,"","","",rack,pos,""}
    return [barcode, "", "", 1, "", "", "", rack, pos, ""]


def format_csv_rows(rows: Sequence[Sequence[Any]], quote: bool = False) -> str:
    """Format rows as a CSV text block.

    - For most PicklyPy outputs, quoting is not used (WL used raw StringRiffle).
    - For merge tables, quoting should be enabled.
    """
    lines: List[str] = []
    for row in rows:
        if quote:
            cells = [quote_csv_field(x) for x in row]
        else:
            cells = [as_str(x) for x in row]
        lines.append(",".join(cells))
    return "\n".join(lines)


# -----------------------------
# Error handling / CLI helpers
# -----------------------------


@dataclass
class RunResult:
    warnings: bool = False
    outputs: List[Path] = None

    def __post_init__(self) -> None:
        if self.outputs is None:
            self.outputs = []


def run_with_user_facing_errors(func, *, pause_on_exit: bool = True) -> int:
    """Run func(), printing user-facing errors and optionally pausing like wolframscript."""
    try:
        func()
        if pause_on_exit:
            try:
                input("Press Enter to exit...")
            except EOFError:
                pass
        return 0
    except PicklyPyError as e:
        print(str(e))
        if pause_on_exit:
            try:
                input("Press Enter to exit...")
            except EOFError:
                pass
        return 2
    except Exception as e:
        print("Unexpected error:")
        traceback.print_exc()
        if pause_on_exit:
            try:
                input("Press Enter to exit...")
            except EOFError:
                pass
        return 1
