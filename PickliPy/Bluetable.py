from __future__ import annotations

"""Blue table picklist generator (Python translation of bluetable.wls).

This script reads:
  1) a source inventory Excel (SRC) file (same structure as PickliPy.Assay SRC)
  2) a folder containing multiple destination “blue table” XLSX files

Each destination XLSX (first worksheet) defines one destination plate; the
destination barcode is taken from the XLSX filename (stem).

All dispenses across all destination files are merged into a single Echo picklist
CSV, plus Plate:Works inventory/process helper files.

The implementation follows the behavior of the original Wolfram Language script
(`bluetable.wls`) while adding validations present in the newer `assay.wls`
lineage (e.g., duplicate compound concentration checks and zero-volume checks).
"""

import argparse
import os
import re
from collections import OrderedDict, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, DefaultDict, Dict, Iterable, List, Mapping, MutableMapping, Optional, Sequence, Tuple

import openpyxl

from common import (
    PicklyPyConfigError,
    PicklyPyError,
    PicklyPyVolumeError,
    RunResult,
    as_number,
    as_str,
    integer_chop,
    load_workbook,
    load_survey_volumes,
    make_inventory_row,
    parse_plate_survey_xml,
    parse_well_name,
    reorder_destinations_within_group,
    reorder_source_groups,
    round_to_increment,
    run_with_user_facing_errors,
    sheet_to_matrix,
    well_name,
    write_text,
)


# -----------------------------
# Constants / layout
# -----------------------------


UNIT_FACTORS: Mapping[str, float] = {
    "nm": 0.001,
    "um": 1.0,
    "mm": 1000.0,
    "ug/ml": 1.0,
    "mg/ml": 1000.0,
    "x": 1.0,
    "": 1.0,
}


# Column groups (Excel 1-based in WL) -> 0-based here.
BLUETABLE_COLS_0BASED: Sequence[Sequence[int]] = (
    (0, 1, 2, 3, 4),
    (6, 7, 8, 9, 10),
    (12, 13, 14, 15, 16),
)


# Destination well offsets per addition block (row_offset, col_offset)
# (WL used {{1,0},{1,4},{1,8}}).
DST_ADDITION_OFFSETS: Sequence[Tuple[int, int]] = (
    (1, 0),
    (1, 4),
    (1, 8),
)


PLATE_ROWS_96 = 8
PLATE_COLS_96 = 12


# -----------------------------
# Data models
# -----------------------------


@dataclass(frozen=True)
class SrcRow:
    """One inventory (SRC) row after normalization."""

    src_well: str  # e.g. A1
    compound_base: str  # e.g. FCCP
    stock_raw: float  # numeric, as entered
    unit_raw: str  # e.g. uM
    stock_base: float  # converted to the “base” unit used in WL (uM-like)
    compound_label: str  # compound_base + stock_base (chopped)
    volume_ul: float
    plate_barcode: str
    vehicle: str
    is_new: bool

    @property
    def src_key(self) -> str:
        # Matches WL convention: "BARCODE_WELL".
        return f"{self.plate_barcode}_{self.src_well}"


@dataclass
class SrcInventory:
    """Parsed SRC inventory and runtime volume state."""

    rows: List[SrcRow]
    min_volume_ul: float
    # mutable volume tracking (keyed by src_key)
    remaining_ul: MutableMapping[str, float]
    # key -> compound_label
    src_name: Mapping[str, str]
    # key -> vehicle type
    vehicles: Mapping[str, str]
    # compound_label -> (list of src_keys, stock_base)
    compound_to_sources: Mapping[str, Tuple[List[str], float]]
    source_plate_barcodes: List[str]
    is_ldv: bool


# -----------------------------
# Formatting helpers
# -----------------------------


def _unit_factor(unit: str) -> float:
    u = as_str(unit).strip().lower()
    if u in UNIT_FACTORS:
        return float(UNIT_FACTORS[u])
    raise PicklyPyConfigError(
        f"Unknown unit {unit!r}. Supported: {sorted(k for k in UNIT_FACTORS if k != '')} (and blank)."
    )


def _fmt_float(x: Any, *, max_decimals: int = 6) -> str:
    """Human-friendly float formatting (avoid 0.30000000004)."""
    if x is None:
        return ""
    try:
        v = float(x)
    except Exception:
        return as_str(x)

    v2 = float(integer_chop(v))
    if abs(v2 - round(v2)) < 1e-12:
        return str(int(round(v2)))
    s = f"{v2:.{max_decimals}f}".rstrip("0").rstrip(".")
    return s


def _cell_table_str(rows: Sequence[Sequence[Any]]) -> str:
    """Format a small 2-column table inside a single well cell."""
    if not rows:
        return "-"
    # Normalize to list of 2-tuples/2-lists for display.
    rows2: List[Tuple[str, str]] = []
    for r in rows:
        if not r:
            continue
        if len(r) == 1:
            rows2.append((as_str(r[0]), ""))
        else:
            rows2.append((as_str(r[0]), _fmt_float(r[1])))
    if not rows2:
        return "-"
    w0 = max(len(a) for a, _ in rows2)
    w1 = max(len(b) for _, b in rows2)
    lines = [f"{a:<{w0}}  {b:>{w1}}".rstrip() for a, b in rows2]
    return "\n".join(lines)


def print_plate_grid(title: str, cells: Sequence[Sequence[str]]) -> None:
    """Print an 8x12 plate grid with ASCII frame and multi-line cells."""
    if len(cells) != PLATE_ROWS_96 or any(len(r) != PLATE_COLS_96 for r in cells):
        raise ValueError("cells must be 8x12")

    col_labels = [str(i) for i in range(1, PLATE_COLS_96 + 1)]
    row_labels = [chr(ord("A") + i) for i in range(PLATE_ROWS_96)]

    # Split cells into lines.
    cell_lines: List[List[List[str]]] = [
        [[(c or "-").splitlines() for c in row] for row in cells]
    ][0]

    col_widths: List[int] = []
    for c in range(PLATE_COLS_96):
        w = len(col_labels[c])
        for r in range(PLATE_ROWS_96):
            for line in cell_lines[r][c]:
                w = max(w, len(line))
        col_widths.append(w)

    row_label_w = 1
    # Horizontal border segments.
    def border() -> str:
        segs = ["+" + "-" * (row_label_w + 2)]
        for w in col_widths:
            segs.append("+" + "-" * (w + 2))
        segs.append("+")
        return "".join(segs)

    def format_row_line(label: str, line_cells: Sequence[str]) -> str:
        out = [f"| {label:<{row_label_w}} "]
        for j, txt in enumerate(line_cells):
            out.append(f"| {txt:<{col_widths[j]}} ")
        out.append("|")
        return "".join(out)

    print(title)
    print(border())
    # Header row
    print(format_row_line("", col_labels))
    print(border())

    for r in range(PLATE_ROWS_96):
        h = max(len(cell_lines[r][c]) for c in range(PLATE_COLS_96))
        for k in range(h):
            line_cells = []
            for c in range(PLATE_COLS_96):
                lines = cell_lines[r][c]
                line_cells.append(lines[k] if k < len(lines) else "")
            print(format_row_line(row_labels[r] if k == 0 else "", line_cells))
        print(border())


# -----------------------------
# Inventory parsing
# -----------------------------


def _detect_minvol_and_columns(src_matrix: List[List[Any]]) -> Tuple[int, int, Optional[int]]:
    """Return (minvol_col_idx0, vehicle_col_idx0, new_col_idx0).

    The historical WL scripts differ slightly:
      - assay.wls: minvol in column G (index 6), "new" marker in col G
      - bluetable.wls: minvol in column H (index 7), vehicle in col G, "new" in col H

    We support both by inspecting row 1.
    """
    if not src_matrix:
        raise PicklyPyConfigError("Empty SRC inventory worksheet.")
    header = src_matrix[0]

    # Candidate columns for minvol: H then G.
    candidates = [7, 6]
    minvol_col = None
    for idx in candidates:
        if idx < len(header) and as_number(header[idx]) is not None:
            minvol_col = idx
            break
    if minvol_col is None:
        raise PicklyPyConfigError(
            "Error: missing minimal Echo Source well volume value in SRC worksheet (expected G1 or H1)."
        )

    # If minvol is in H, assume G is vehicle and H is new-marker.
    if minvol_col == 7:
        vehicle_col = 6
        new_col = 7
    else:
        # If minvol is in G, there is no vehicle column (assay-style); new marker is in G.
        vehicle_col = -1
        new_col = 6

    return minvol_col, vehicle_col, new_col


def _load_src_inventory(inventory_xlsx: Path, folderdst: Path, *, out_dir: Path) -> SrcInventory:
    if not inventory_xlsx.exists():
        raise PicklyPyConfigError(f"Inventory file not found: {inventory_xlsx}")

    wb = load_workbook(inventory_xlsx, data_only=True)
    ws = wb.worksheets[0]
    src_matrix = sheet_to_matrix(ws, min_cols=8)

    minvol_col, vehicle_col, new_col = _detect_minvol_and_columns(src_matrix)
    minvol = as_number(src_matrix[0][minvol_col])
    if minvol is None or not (minvol >= 0):
        raise PicklyPyConfigError(
            "Error: missing minimal Echo Source well volume value in SRC worksheet (G1/H1)."
        )

    # Parse raw rows.
    raw_rows: List[Dict[str, Any]] = []
    for r in src_matrix[1:]:
        # Ensure row is long enough.
        if len(r) < 6:
            continue
        src_well = as_str(r[0]).strip()
        compound = as_str(r[1]).strip()
        stock_raw = as_number(r[2])
        unit_raw = as_str(r[3]).strip()
        vol_ul = as_number(r[4])
        barcode = as_str(r[5]).strip()

        if src_well == "" and barcode == "" and compound == "":
            continue
        if src_well == "" or barcode == "":
            # WL scripts effectively skip blank well/barcode lines.
            continue
        if compound == "":
            raise PicklyPyConfigError(
                f"SRC row with Plate barcode={barcode!r} Well={src_well!r} is missing Compound Name."
            )
        if stock_raw is None:
            raise PicklyPyConfigError(
                f"SRC row for {barcode}_{src_well} ({compound}) is missing numeric Stock concentration."
            )
        if vol_ul is None:
            raise PicklyPyConfigError(
                f"SRC row for {barcode}_{src_well} ({compound}) is missing numeric Volume in source plate (uL)."
            )

        vehicle = ""
        if vehicle_col >= 0 and vehicle_col < len(r):
            vehicle = as_str(r[vehicle_col]).strip()
        is_new = False
        if new_col < len(r):
            is_new = as_str(r[new_col]).strip().lower() == "new"

        raw_rows.append(
            {
                "src_well": src_well,
                "compound": compound,
                "stock_raw": float(stock_raw),
                "unit_raw": unit_raw,
                "volume_ul": float(vol_ul),
                "barcode": barcode,
                "vehicle": vehicle,
                "is_new": is_new,
            }
        )

    if not raw_rows:
        raise PicklyPyConfigError("No usable SRC rows found in the inventory file.")

    # Update volumes from survey XMLs found in folderdst.
    # Common helper loads all *_platesurvey.xml files in a folder.
    survey = load_survey_volumes(folderdst)
    survey_vols = dict(survey.volumes_ul)

    # Handle UnknownBarCode the same way assay.wls does.
    unique_barcodes = sorted({rr["barcode"] for rr in raw_rows if rr["barcode"] != ""})
    if "UnknownBarCode" in {bc for bc, _ in survey_vols.keys()}:
        if len(unique_barcodes) == 1:
            fallback = unique_barcodes[0]
            remapped: Dict[Tuple[str, str], float] = {}
            for (bc, well), vol in survey_vols.items():
                if bc == "UnknownBarCode":
                    remapped[(fallback, well)] = vol
                else:
                    remapped[(bc, well)] = vol
            survey_vols = remapped
        else:
            raise PicklyPyConfigError(
                "Plate survey file barcode is 'UnknownBarCode' but the inventory uses multiple source barcodes; "
                "cannot disambiguate."
            )

    if survey_vols:
        missing: List[str] = []
        for rr in raw_rows:
            key = (rr["barcode"], rr["src_well"])
            if key in survey_vols:
                rr["volume_ul"] = float(survey_vols[key])
            else:
                # Match WL behavior: set to 0 (and warn) when survey exists but value is missing.
                missing.append(f"{rr['barcode']}_{rr['src_well']}")
                rr["volume_ul"] = 0.0
        if missing:
            preview = ", ".join(missing[:12]) + (" ..." if len(missing) > 12 else "")
            print(
                "Warning: Using survey file(s) but not finding values for one or more wells; using 0 uL for these. "
                f"Missing count={len(missing)}. Examples: {preview}"
            )

        # Restore volumes for wells marked 'new'.
        # (WL bluetable keeps original volume for those.)
        # We need the original values from the workbook.
        # Re-read original volumes from src_matrix for those rows in order.
        orig_vol_by_key: Dict[Tuple[str, str], float] = {}
        for r in src_matrix[1:]:
            if len(r) < 6:
                continue
            w = as_str(r[0]).strip()
            bc = as_str(r[5]).strip()
            if w == "" or bc == "":
                continue
            v = as_number(r[4])
            if v is None:
                continue
            orig_vol_by_key[(bc, w)] = float(v)
        for rr in raw_rows:
            if rr["is_new"]:
                k = (rr["barcode"], rr["src_well"])
                if k in orig_vol_by_key:
                    rr["volume_ul"] = orig_vol_by_key[k]

        # Export a copy of the inventory after survey update.
        after_survey_path = out_dir / f"after_survey_orig_{inventory_xlsx.name}"
        _export_src_like_workbook(after_survey_path, src_matrix, raw_rows, volume_col_idx0=4)
        print(f"Volumes updated from survey file(s). Saved: {after_survey_path}")

    # Validations (ported from assay.wls and bluetable.wls)
    # 1) Source well + plate barcode must be unique.
    seen: set[Tuple[str, str]] = set()
    dups: List[str] = []
    for rr in raw_rows:
        k = (rr["barcode"], rr["src_well"])
        if k in seen:
            dups.append(f"{k[0]}_{k[1]}")
        else:
            seen.add(k)
    if dups:
        raise PicklyPyConfigError(
            "Error: duplicate well assignment in SRC worksheet within the same source plate: "
            + ", ".join(sorted(set(dups)))
        )

    # 2) Duplicate compound labels must have the same stock concentration within each source plate.
    # (This validation was present in assay.wls; bluetable.wls had the data prep but missing the check.)
    by_plate_compound: DefaultDict[Tuple[str, str], set[float]] = defaultdict(set)
    for rr in raw_rows:
        try:
            f = _unit_factor(rr["unit_raw"])
        except PicklyPyConfigError:
            # give a more specific context
            raise PicklyPyConfigError(
                f"Unknown unit {rr['unit_raw']!r} in SRC row for {rr['barcode']}_{rr['src_well']} ({rr['compound']})."
            )
        by_plate_compound[(rr["barcode"], rr["compound"])].add(float(rr["stock_raw"]) * f)
    bad = [k for k, concs in by_plate_compound.items() if len(concs) > 1]
    if bad:
        preview = ", ".join([f"{plate}:{comp}" for plate, comp in bad[:10]]) + (" ..." if len(bad) > 10 else "")
        raise PicklyPyConfigError(
            "Error: duplicate compound labels in SRC worksheet must have the same concentrations within each source plate. "
            f"Offenders: {preview}"
        )

    # Normalize rows to WL behavior: convert stock to base units and append to compound name.
    rows: List[SrcRow] = []
    for rr in raw_rows:
        f = _unit_factor(rr["unit_raw"])
        stock_base = float(rr["stock_raw"]) * f
        stock_base = float(integer_chop(stock_base))
        compound_label = f"{rr['compound']}{as_str(integer_chop(stock_base))}"
        rows.append(
            SrcRow(
                src_well=rr["src_well"],
                compound_base=rr["compound"],
                stock_raw=float(rr["stock_raw"]),
                unit_raw=rr["unit_raw"],
                stock_base=stock_base,
                compound_label=compound_label,
                volume_ul=float(rr["volume_ul"]),
                plate_barcode=rr["barcode"],
                vehicle=rr["vehicle"],
                is_new=bool(rr["is_new"]),
            )
        )

    remaining_ul: Dict[str, float] = OrderedDict((r.src_key, r.volume_ul) for r in rows)
    src_name: Dict[str, str] = {r.src_key: r.compound_label for r in rows}
    vehicles: Dict[str, str] = {r.src_key: r.vehicle for r in rows}
    source_plate_barcodes = sorted({r.plate_barcode for r in rows if r.plate_barcode != ""})

    # Pool alternative wells for identical compounds (compound_label is already name+stock).
    compound_to_sources: Dict[str, Tuple[List[str], float]] = {}
    tmp: DefaultDict[str, List[SrcRow]] = defaultdict(list)
    for r in rows:
        tmp[r.compound_label].append(r)
    for comp, comp_rows in tmp.items():
        comp_rows_sorted = sorted(comp_rows, key=lambda x: (x.plate_barcode, x.src_well))
        compound_to_sources[comp] = ([r.src_key for r in comp_rows_sorted], comp_rows_sorted[0].stock_base)

    # LDV detection (WL: max volume < 15 uL => LDV).
    is_ldv = max(r.volume_ul for r in rows) < 15.0

    print(f"{inventory_xlsx} loaded successfully.")
    return SrcInventory(
        rows=rows,
        min_volume_ul=float(minvol),
        remaining_ul=remaining_ul,
        src_name=src_name,
        vehicles=vehicles,
        compound_to_sources=compound_to_sources,
        source_plate_barcodes=source_plate_barcodes,
        is_ldv=is_ldv,
    )


def _export_src_like_workbook(
    out_path: Path,
    src_matrix_original: List[List[Any]],
    raw_rows_current: List[Dict[str, Any]],
    *,
    volume_col_idx0: int,
) -> None:
    """Write a copy of the SRC-like sheet, updating only the volume column.

    The WL scripts export the full table after survey/dispense as an Excel file.
    We approximate this by copying the original matrix and updating the volume
    column for those rows we parsed.
    """

    # Build lookup by (barcode, well) -> volume_ul
    vol_by_key: Dict[Tuple[str, str], float] = {
        (rr["barcode"], rr["src_well"]): float(rr["volume_ul"]) for rr in raw_rows_current
    }
    mat = [list(r) for r in src_matrix_original]
    # Update row-by-row for rows that match.
    for i in range(1, len(mat)):
        row = mat[i]
        if len(row) < 6:
            continue
        w = as_str(row[0]).strip()
        bc = as_str(row[5]).strip()
        if w == "" or bc == "":
            continue
        k = (bc, w)
        if k in vol_by_key:
            # Ensure row long enough.
            while len(row) <= volume_col_idx0:
                row.append("")
            row[volume_col_idx0] = vol_by_key[k]
    # Write workbook.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SRC"
    for r_idx, row in enumerate(mat, start=1):
        for c_idx, v in enumerate(row, start=1):
            ws.cell(r_idx, c_idx).value = v
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


# -----------------------------
# Blue table parsing / dispensing
# -----------------------------


_WELL_RE = re.compile(r"^[A-Za-z]+\d+$")


def _is_well_name(s: str) -> bool:
    return bool(_WELL_RE.fullmatch(s.strip()))


def _split_treatments(treat: str) -> List[str]:
    # Match WL separators: + , ; and newlines.
    parts = re.split(r"[\+\,;\n\r]+", as_str(treat))
    return [p.strip() for p in parts if p.strip() != ""]


def _fix_unit_row(
    *,
    dst_well: str,
    treat_name: str,
    final_raw: Any,
    stock_raw: Any,
    unit_raw: Any,
) -> Tuple[str, str, float]:
    """WL fixunit: {well, treat<>stockBase, finalBase}."""
    f = _unit_factor(as_str(unit_raw))
    final = as_number(final_raw)
    stock = as_number(stock_raw)
    if final is None or stock is None:
        raise PicklyPyConfigError(
            f"Missing numeric Final/Stock for treatment {treat_name!r} in destination well {dst_well!r}."
        )
    stock_base = float(integer_chop(float(stock) * f))
    treat_label = f"{as_str(treat_name)}{as_str(integer_chop(stock_base))}"
    final_base = float(final) * f
    return dst_well, treat_label, final_base


def _calc_echo_volume_nl(
    *,
    stock_base: float,
    final_base: float,
    wellvolume_ul: float,
    concentrated: float,
    rounding_inc_nl: float = 2.5,
) -> float:
    # WL: Round[concentrated*1000*wellvolume*final/stock, 2.5]
    if stock_base <= 0:
        raise PicklyPyConfigError(f"Invalid stock concentration (<=0): {stock_base}")
    desired_nl = float(concentrated) * 1000.0 * float(wellvolume_ul) * float(final_base) / float(stock_base)
    rounded = float(round_to_increment(desired_nl, rounding_inc_nl))
    # Validation added from assay.wls: forbid zero-volume dispenses.
    if desired_nl > 0 and rounded == 0:
        raise PicklyPyConfigError(
            f"Error: Zero-volume dispensing after rounding. Desired={desired_nl} nL -> Rounded={rounded} nL."
        )
    return rounded


def _warn_rounding(
    *,
    warnings_flag: List[bool],
    compound_label: str,
    stock_base: float,
    final_base: float,
    wellvolume_ul: float,
    concentrated: float,
    rounded_nl: float,
    threshold: float = 0.2,
) -> None:
    # WL bluetable used 0.2.
    desired_nl = float(concentrated) * 1000.0 * float(wellvolume_ul) * float(final_base) / float(stock_base)
    if desired_nl <= 0:
        return
    rel = abs(desired_nl - rounded_nl) / desired_nl
    if rel > threshold:
        warnings_flag[0] = True
        possible = max(2.5, rounded_nl)
        print(
            "Warning: Significant rounding error at low volume dispensing! "
            f"Compound: {compound_label} Stock(base): {stock_base} Final(base): {final_base} "
            f"Desired: {desired_nl} nL Possible: {possible} nL"
        )


def _choose_src_well(
    *,
    src_candidates: Sequence[str],
    vol_nl: float,
    inv: SrcInventory,
) -> str:
    need_ul = float(vol_nl) / 1000.0
    for k in src_candidates:
        have_ul = float(inv.remaining_ul.get(k, 0.0))
        if have_ul >= float(inv.min_volume_ul) + need_ul:
            inv.remaining_ul[k] = have_ul - need_ul
            return k
    # Out of volume.
    first = src_candidates[0] if src_candidates else "(none)"
    comp = inv.src_name.get(first, first)
    raise PicklyPyVolumeError(f"Error: Out of Volume for {comp}")


def _grid_index_96(well: str) -> Tuple[int, int]:
    r, c = parse_well_name(well)
    if not (1 <= r <= PLATE_ROWS_96 and 1 <= c <= PLATE_COLS_96):
        raise PicklyPyConfigError(
            f"Destination well {well!r} is outside the 96-well display grid (A-H,1-12). "
            "Check offsets / input well labels."
        )
    return r - 1, c - 1


def _read_cell(ws: openpyxl.worksheet.worksheet.Worksheet, addr: str, *, numeric: bool = False) -> Any:
    v = ws[addr].value
    return as_number(v) if numeric else v


def _check_well_label_consistency(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    *,
    block_starts_1based: Sequence[int] = (1, 7, 13),
    name_offset: int = 1,
    row_min: int = 2,
    row_max: int = 25,
) -> None:
    """Port of the WL consistency check.

    Ensures that the per-row well label is consistent across addition blocks
    (ignoring rows where the name column is blank, and ignoring def/veh rows).
    """
    bad_rows: List[int] = []
    for r in range(row_min, row_max + 1):
        wells: List[str] = []
        for c0 in block_starts_1based:
            w = as_str(ws.cell(r, c0).value).strip()
            name = as_str(ws.cell(r, c0 + name_offset).value).strip()
            if name == "":
                continue
            if w.lower() in {"def", "veh"}:
                continue
            if w == "" or w == "0":
                continue
            wells.append(w)
        if len(set(wells)) > 1:
            bad_rows.append(r)
    if bad_rows:
        raise PicklyPyConfigError(
            "Error: Mismatching well name definitions in destination file rows: "
            + ", ".join(str(r) for r in bad_rows)
        )


def _process_one_bluetable_file(
    dst_xlsx: Path,
    *,
    inv: SrcInventory,
    picklist_rows: List[List[Any]],
    total_use_rows: List[Tuple[str, str, float]],
    manual_labels: set[str],
    warnings_flag: List[bool],
    tables_out_dir: Path,
) -> None:
    if not dst_xlsx.exists():
        raise PicklyPyConfigError(f"Destination file not found: {dst_xlsx}")

    wb = load_workbook(dst_xlsx, data_only=True)
    ws = wb.worksheets[0]
    dst_barcode = dst_xlsx.stem
    print(f"****** Barcode: {dst_barcode} ******")

    # Check row-by-row well label consistency across blocks.
    _check_well_label_consistency(ws)

    # Read per-block parameters.
    wellvolumes = [
        _read_cell(ws, "C29", numeric=True),
        _read_cell(ws, "E29", numeric=True),
        _read_cell(ws, "G29", numeric=True),
    ]
    prepvolumes = [
        _read_cell(ws, "C31", numeric=True),
        _read_cell(ws, "E31", numeric=True),
        _read_cell(ws, "G31", numeric=True),
    ]
    dilution = _read_cell(ws, "I28", numeric=True)
    if dilution is None:
        dilution = 1.0
    concentrated = float(dilution)

    # Manual and Echo tables are 8x12 (as in WL notebook grid).
    manual_table_raw: List[List[List[List[Any]]]] = [
        [[[] for _ in range(PLATE_COLS_96)] for _ in range(PLATE_ROWS_96)]
    ][0]

    # Collect all dispense events (dst_well, src_candidates, vol_nl).
    all_events: List[Tuple[str, List[str], float]] = []

    # Determine how many blocks are present/used.
    for block_idx, cols0 in enumerate(BLUETABLE_COLS_0BASED):
        if block_idx >= len(DST_ADDITION_OFFSETS):
            continue
        offset_r, offset_c = DST_ADDITION_OFFSETS[block_idx]
        wellvol = wellvolumes[block_idx] if block_idx < len(wellvolumes) else None
        prepvol = prepvolumes[block_idx] if block_idx < len(prepvolumes) else None
        # Extract rows (Excel 2..25 inclusive).
        rows_block: List[List[Any]] = []
        for r in range(2, 26):
            row = [ws.cell(r, c0 + 1).value for c0 in cols0]  # openpyxl is 1-based
            rows_block.append(row)

        # Remove medium rows (WL: DeleteCases[a,_?(#[[2]]=="medium"&)].)
        rows_block = [
            row
            for row in rows_block
            if as_str(row[1]).strip().lower() != "medium"
        ]

        # If block has no content, skip silently.
        has_any = any(as_str(row[1]).strip() != "" for row in rows_block)
        if not has_any:
            continue
        if wellvol is None or wellvol <= 0:
            raise PicklyPyConfigError(
                f"Destination file {dst_xlsx.name}: block #{block_idx+1} has entries but Well volume is missing/zero."
            )
        if prepvol is None or prepvol < 0:
            raise PicklyPyConfigError(
                f"Destination file {dst_xlsx.name}: block #{block_idx+1} has entries but Prep volume is missing/invalid."
            )

        wellvolume_ul = float(wellvol)
        prepvolume_ul = float(prepvol)

        # Build defs: treat_name -> (treat_name, final, stock, unit)
        defs: Dict[str, Tuple[str, Any, Any, Any]] = {}
        veh_rows: List[Tuple[str, str, Any, Any, Any]] = []

        for row in rows_block:
            kind = as_str(row[0]).strip().lower()
            name = as_str(row[1]).strip()
            if kind in {"def", "veh"} and name != "":
                defs[name] = (name, row[2], row[3], row[4])
            if kind == "veh" and name != "":
                # WL: vehs[[All,1]] = vehs[[All,2]]
                veh_rows.append((name, name, row[2], row[3], row[4]))

        # Collect normal dispense rows.
        dispense_rows: List[Tuple[str, str, Any, Any, Any]] = []
        for row in rows_block:
            w_raw = as_str(row[0]).strip()
            treat_raw = as_str(row[1]).strip()
            if treat_raw == "":
                continue
            if w_raw.lower() in {"def", "veh"} or w_raw == "":
                continue
            if not _is_well_name(w_raw):
                # WL required string well names; skip obvious placeholders like 0.
                continue
            dispense_rows.append((w_raw, treat_raw, row[2], row[3], row[4]))

        # Apply destination offsets.
        shifted_rows: List[Tuple[str, str, Any, Any, Any]] = []
        for w_raw, treat_raw, fin, stock, unit in dispense_rows:
            rr, cc = parse_well_name(w_raw)
            dst_well = well_name(rr + offset_r, cc + offset_c)
            shifted_rows.append((dst_well, treat_raw, fin, stock, unit))

        # Expand combined treatments.
        expanded: List[Tuple[str, str, Any, Any, Any]] = []
        missing_defs: set[str] = set()
        for dst_well, treat_raw, fin, stock, unit in shifted_rows:
            parts = _split_treatments(treat_raw)
            if not parts:
                continue
            if len(parts) == 1:
                expanded.append((dst_well, parts[0], fin, stock, unit))
                continue
            # All but last must come from defs.
            for p in parts[:-1]:
                if p not in defs:
                    missing_defs.add(p)
                else:
                    name, dfin, dstock, dunit = defs[p]
                    expanded.append((dst_well, name, dfin, dstock, dunit))
            # Last uses the row values.
            expanded.append((dst_well, parts[-1], fin, stock, unit))

        if missing_defs:
            raise PicklyPyConfigError(
                "Undefined combined addition component(s): "
                + ", ".join(sorted(missing_defs))
                + "\nUse 'def' lines to define these components."
            )

        # Convert to (dst_well, treat_label, final_base).
        fixed = [
            _fix_unit_row(dst_well=dw, treat_name=tn, final_raw=fin, stock_raw=st, unit_raw=un)
            for (dw, tn, fin, st, un) in expanded
        ]

        # Map to inventory sources (or manual dilutions if missing).
        cecho: List[Tuple[str, List[str], float, float]] = []  # dst, src_candidates, stock_base, final_base
        cmanual: List[Tuple[str, str, float]] = []  # dst, treat_label, final_base
        for dst_well, treat_label, final_base in fixed:
            if treat_label in inv.compound_to_sources:
                src_candidates, stock_base = inv.compound_to_sources[treat_label]
                cecho.append((dst_well, src_candidates, float(stock_base), float(final_base)))
            else:
                cmanual.append((dst_well, treat_label, float(final_base)))

        print(f"Dispensing into: {dst_barcode} (block #{block_idx+1})")

        # Manual dilutions (not in inventory)
        if cmanual:
            manual_list = []
            for dst_well, name_label, final_base in cmanual:
                md = _manual_dilute(
                    name_label=name_label,
                    final_base=final_base,
                    wellvolume_ul=wellvolume_ul,
                    prepvolume_ul=prepvolume_ul,
                    concentrated=concentrated,
                )
                manual_list.append((dst_well, md))
            print("Manual dilution (not in inventory):")
            for dst_well, (name, vol_stock, vol_medium) in manual_list:
                print(f"  {dst_well}: {name}  stock={_fmt_float(vol_stock)} uL  medium={_fmt_float(vol_medium)} uL")
            manual_labels.update({name_label for _, name_label, _ in cmanual})

            # Populate manual table (raw 3-column rows).
            for dst_well, name_label, final_base in cmanual:
                i, j = _grid_index_96(dst_well)
                manual_table_raw[i][j].append(
                    list(
                        _manual_dilute(
                            name_label=name_label,
                            final_base=final_base,
                            wellvolume_ul=wellvolume_ul,
                            prepvolume_ul=prepvolume_ul,
                            concentrated=concentrated,
                        )
                    )
                )

        # For wells that have Echo additions but no manual additions, set Medium=prepvolume.
        for dst_well, _, _, _ in cecho:
            i, j = _grid_index_96(dst_well)
            if manual_table_raw[i][j] == []:
                manual_table_raw[i][j] = [["Medium", prepvolume_ul]]

        # Echo events for compounds
        l0: List[Tuple[str, List[str], float]] = []
        for dst_well, src_candidates, stock_base, final_base in cecho:
            vol_nl = _calc_echo_volume_nl(
                stock_base=stock_base,
                final_base=final_base,
                wellvolume_ul=wellvolume_ul,
                concentrated=concentrated,
            )
            # Rounding warning
            # Use the compound label based on the first candidate.
            compound_label = inv.src_name.get(src_candidates[0], src_candidates[0])
            _warn_rounding(
                warnings_flag=warnings_flag,
                compound_label=compound_label,
                stock_base=stock_base,
                final_base=final_base,
                wellvolume_ul=wellvolume_ul,
                concentrated=concentrated,
                rounded_nl=vol_nl,
            )
            l0.append((dst_well, list(src_candidates), float(vol_nl)))

        # Vehicle definitions (veh rows)
        vehicle_defs: List[Tuple[str, List[str], float]] = []  # (vehicle_type, src_candidates, required_vol_nl)
        for veh_type, _, fin, stock, unit in veh_rows:
            # Convert veh row to treat label, then look up in inventory.
            _dw, treat_label, final_base = _fix_unit_row(
                dst_well=veh_type,
                treat_name=veh_type,
                final_raw=fin,
                stock_raw=stock,
                unit_raw=unit,
            )
            if treat_label not in inv.compound_to_sources:
                raise PicklyPyConfigError(
                    f"Vehicle '{veh_type}' (label {treat_label}) is not found in inventory SRC; cannot auto-add vehicle."
                )
            src_candidates, stock_base = inv.compound_to_sources[treat_label]
            req_nl = _calc_echo_volume_nl(
                stock_base=float(stock_base),
                final_base=float(final_base),
                wellvolume_ul=wellvolume_ul,
                concentrated=concentrated,
            )
            vehicle_defs.append((veh_type, list(src_candidates), float(req_nl)))

        # Auto-add vehicle to wells that contain vehicle-containing additions.
        if vehicle_defs:
            # For each l0 dispense, map it to its vehicle type (based on first src well candidate).
            per_vehicle_per_well: DefaultDict[Tuple[str, str], float] = defaultdict(float)
            wells_with_vehicle: DefaultDict[str, set[str]] = defaultdict(set)
            for dst_well, src_candidates, vol_nl in l0:
                src0 = src_candidates[0]
                vtype = inv.vehicles.get(src0, "")
                if vtype == "":
                    continue
                per_vehicle_per_well[(vtype, dst_well)] += float(vol_nl)
                wells_with_vehicle[vtype].add(dst_well)

            for vtype, v_src_candidates, req_nl in vehicle_defs:
                # Only add to wells that have at least one compound using this vehicle type.
                for dst_well in sorted(wells_with_vehicle.get(vtype, set()), key=lambda w: parse_well_name(w)):
                    have_nl = per_vehicle_per_well.get((vtype, dst_well), 0.0)
                    if have_nl + 1e-9 < req_nl:
                        vehicle_add_nl = req_nl - have_nl
                        l0.append((dst_well, list(v_src_candidates), float(vehicle_add_nl)))

        # Append to overall event list.
        all_events.extend(l0)

    # If nothing to dispense from this destination plate, skip but still generate a tables file.
    if not all_events:
        print(f"No dispense events found in {dst_xlsx.name}. Skipping picklist lines.")

    # Group events by compound (src_candidates list identity), pool duplicates by destination well.
    by_compound: DefaultDict[Tuple[str, ...], Dict[str, float]] = defaultdict(lambda: defaultdict(float))
    for dst_well, src_candidates, vol_nl in all_events:
        key = tuple(src_candidates)
        by_compound[key][dst_well] += float(vol_nl)

    groups: List[List[Tuple[str, List[str], float]]] = []
    for src_candidates_key, dst_to_vol in by_compound.items():
        src_candidates = list(src_candidates_key)
        events = [(dst_w, src_candidates, float(vol)) for dst_w, vol in dst_to_vol.items()]
        groups.append(events)

    # Reorder groups and destinations (travel optimization).
    groups = reorder_source_groups(groups)
    last_pos = (1, 1)
    ordered_events: List[Tuple[str, List[str], float]] = []
    for g in groups:
        g2, last_pos = reorder_destinations_within_group(g, last_pos)
        ordered_events.extend(g2)

    # Choose concrete source wells and update remaining volumes.
    chosen: List[Tuple[str, str, float]] = []  # dst, src_key, vol_nl
    for dst_well, src_candidates, vol_nl in ordered_events:
        src_key = _choose_src_well(src_candidates=src_candidates, vol_nl=vol_nl, inv=inv)
        chosen.append((dst_well, src_key, float(vol_nl)))

    # Group by chosen source well (preserve first-seen order like Mathematica GatherBy).
    by_src: "OrderedDict[str, List[Tuple[float, str]]]" = OrderedDict()
    for dst_well, src_key, vol_nl in chosen:
        if inv.is_ldv and vol_nl > 500.0:
            # Split large volumes into <=500nL chunks (WL behavior for LDV).
            remaining = vol_nl
            while remaining > 0:
                chunk = min(500.0, remaining)
                by_src.setdefault(src_key, []).append((float(chunk), dst_well))
                remaining -= chunk
        else:
            by_src.setdefault(src_key, []).append((float(vol_nl), dst_well))

    # Volume usage summary for this destination barcode.
    print("Volumes used:")
    for src_key, pairs in by_src.items():
        used_ul = sum(v for v, _ in pairs) / 1000.0
        comp = inv.src_name.get(src_key, src_key)
        print(f"  {comp}   {src_key}   {used_ul} uL")
        total_use_rows.append((comp, src_key, float(used_ul)))

    # Append picklist CSV rows.
    for src_key, pairs in by_src.items():
        plate_bc, src_well = src_key.rsplit("_", 1)
        row: List[Any] = [plate_bc, src_well, dst_barcode]
        for vol_nl, dst_well in pairs:
            row.append(integer_chop(vol_nl))
            row.append(dst_well)
        picklist_rows.append(row)

    # Build and print plate grids (Manual and Echo)
    # Manual: convert raw 3-col rows into the 2-col display table + Total Medium.
    manual_cells: List[List[str]] = [["-" for _ in range(PLATE_COLS_96)] for _ in range(PLATE_ROWS_96)]
    for i in range(PLATE_ROWS_96):
        for j in range(PLATE_COLS_96):
            cell = manual_table_raw[i][j]
            if cell == []:
                manual_cells[i][j] = "-"
                continue
            # If this is a list of 3-col manual dilutes, compute Total Medium.
            if isinstance(cell, list) and cell and isinstance(cell[0], list) and len(cell[0]) >= 3:
                # Table of (name, vol_stock)
                rows2 = [[r[0], r[1]] for r in cell]
                prep_total = float(cell[0][1]) + float(cell[0][2])
                used_stock = sum(float(r[1]) for r in cell)
                total_medium = prep_total - used_stock
                rows2.append(["Total Medium", total_medium])
                manual_cells[i][j] = _cell_table_str(rows2)
            else:
                manual_cells[i][j] = _cell_table_str(cell)

    print_plate_grid("Manual additions:", manual_cells)

    # Echo: fill from chosen events.
    echo_raw: List[List[List[List[Any]]]] = [
        [[[] for _ in range(PLATE_COLS_96)] for _ in range(PLATE_ROWS_96)]
    ][0]
    for dst_well, src_key, vol_nl in chosen:
        i, j = _grid_index_96(dst_well)
        echo_raw[i][j].append([inv.src_name.get(src_key, src_key), integer_chop(vol_nl)])

    echo_cells: List[List[str]] = [["-" for _ in range(PLATE_COLS_96)] for _ in range(PLATE_ROWS_96)]
    for i in range(PLATE_ROWS_96):
        for j in range(PLATE_COLS_96):
            echo_cells[i][j] = _cell_table_str(echo_raw[i][j])

    print_plate_grid("Echo additions:", echo_cells)

    # Save notebook tables xlsx (Echo + Manual) with headers.
    tables_out_dir.mkdir(parents=True, exist_ok=True)
    out_tables = tables_out_dir / f"{dst_barcode}_notebook_tables.xlsx"
    _write_notebook_tables_xlsx(out_tables, echo_cells, manual_cells)
    print(f"Saving: {out_tables}")


def _manual_dilute(
    *,
    name_label: str,
    final_base: float,
    wellvolume_ul: float,
    prepvolume_ul: float,
    concentrated: float,
) -> Tuple[str, float, float]:
    """WL manualdilute: derive stock concentration from trailing digits and compute volumes.

    Returns (name_label, vol_stock_ul, vol_medium_ul).
    """
    m = re.search(r"(\d+)\s*$", name_label)
    if not m:
        raise PicklyPyConfigError(
            f"Manual dilution compound label {name_label!r} does not end with a numeric stock concentration. "
            "This label should be like 'Compound1234' (stock in base units)."
        )
    stock_base = float(m.group(1))
    if stock_base <= 0:
        raise PicklyPyConfigError(f"Invalid stock concentration parsed from {name_label!r}: {stock_base}")
    vol_stock = float(concentrated) * float(wellvolume_ul) * float(final_base) / stock_base
    vol_medium = float(prepvolume_ul) - vol_stock
    if vol_medium < -1e-9:
        raise PicklyPyConfigError(
            f"Manual dilution negative medium volume for {name_label} (stock vol {vol_stock} > prep {prepvolume_ul})."
        )
    return name_label, vol_stock, vol_medium


def _write_notebook_tables_xlsx(path: Path, echo_cells: Sequence[Sequence[str]], manual_cells: Sequence[Sequence[str]]) -> None:
    """Write the per-destination notebook tables workbook.

    WL produced 2 sheets: Echo and Manual, both with row/col headers.
    """
    wb = openpyxl.Workbook()
    # Remove default
    wb.remove(wb.active)

    def add_sheet(title: str, cells: Sequence[Sequence[str]]) -> None:
        ws = wb.create_sheet(title)
        # Header row
        ws.cell(1, 1).value = 0
        for c in range(PLATE_COLS_96):
            ws.cell(1, 2 + c).value = c + 1
        # Data rows with row labels
        for r in range(PLATE_ROWS_96):
            ws.cell(2 + r, 1).value = chr(ord("A") + r)
            for c in range(PLATE_COLS_96):
                ws.cell(2 + r, 2 + c).value = cells[r][c]

    add_sheet("Echo", echo_cells)
    add_sheet("Manual", manual_cells)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


# -----------------------------
# Main run
# -----------------------------


def run_bluetable(
    *,
    picklist_base: Path,
    inventory_xlsx: Path,
    folderdst: Path,
    tables_out_dir: Optional[Path] = None,
) -> RunResult:
    if not folderdst.exists() or not folderdst.is_dir():
        raise PicklyPyConfigError(f"Destination folder not found: {folderdst}")

    out_dir = picklist_base.parent if picklist_base.parent != Path("") else Path.cwd()
    out_dir.mkdir(parents=True, exist_ok=True)
    tables_out_dir = tables_out_dir or out_dir

    print(
        "Blue table picklist generator for acoustic dispensing using Beckman Coulter Echo 650 controlled by Revvity PlateWorks. "
        "Written by Akos A. Gerencser."
    )
    print("Python translation of bluetable.wls")

    inv = _load_src_inventory(inventory_xlsx, folderdst, out_dir=out_dir)
    if len(inv.source_plate_barcodes) != 1:
        print(
            f"Warning: inventory contains {len(inv.source_plate_barcodes)} source plate barcodes. "
            "BlueTable workflows typically use a single source plate."
        )

    # Destination XLSX files.
    dst_files = sorted(folderdst.glob("*.xlsx"))
    dst_files = [p for p in dst_files if "~$" not in p.name]
    dst_files = [p for p in dst_files if not p.name.endswith("_notebook_tables.xlsx")]
    if not dst_files:
        raise PicklyPyConfigError(f"No destination .xlsx files found in: {folderdst}")

    picklist_rows: List[List[Any]] = []
    total_use_rows: List[Tuple[str, str, float]] = []
    manual_labels: set[str] = set()
    warnings_flag = [False]

    for dst in dst_files:
        _process_one_bluetable_file(
            dst,
            inv=inv,
            picklist_rows=picklist_rows,
            total_use_rows=total_use_rows,
            manual_labels=manual_labels,
            warnings_flag=warnings_flag,
            tables_out_dir=tables_out_dir,
        )

    # Write picklist CSV.
    picklist_csv_path = out_dir / f"{picklist_base.name}_Picklist.csv"
    picklist_text_lines: List[str] = []
    for row in picklist_rows:
        picklist_text_lines.append(",".join(as_str(x) for x in row))
    write_text(picklist_csv_path, "\n".join(picklist_text_lines) + ("\n" if picklist_text_lines else ""))
    print(f"Saving: {picklist_csv_path}")

    # Summary: total usage per source well.
    print("****** Summary: Total Usage ******")
    by_src: DefaultDict[str, List[Tuple[str, float]]] = defaultdict(list)
    for comp, src_key, used_ul in total_use_rows:
        by_src[src_key].append((comp, float(used_ul)))
    for src_key, items in by_src.items():
        comp = items[0][0]
        total_ul = sum(u for _, u in items)
        remains = inv.remaining_ul.get(src_key, 0.0)
        print(f"{comp}\t\t{src_key}\t{total_ul}\tremains (ul): {remains}")

    # Export inventory after dispense.
    # Re-read original SRC matrix to preserve formatting/columns as much as possible.
    wb0 = load_workbook(inventory_xlsx, data_only=True)
    ws0 = wb0.worksheets[0]
    src_matrix0 = sheet_to_matrix(ws0, min_cols=8)
    raw_rows_for_export: List[Dict[str, Any]] = []
    for r in inv.rows:
        raw_rows_for_export.append(
            {
                "src_well": r.src_well,
                "barcode": r.plate_barcode,
                "volume_ul": float(inv.remaining_ul.get(r.src_key, 0.0)),
            }
        )
    # We need the full dict form used by _export_src_like_workbook.
    # Merge back the missing fields as blanks.
    # Build row-keyed dict from inv.rows to keep unique.
    raw_by_key = {(r.plate_barcode, r.src_well): float(inv.remaining_ul.get(r.src_key, 0.0)) for r in inv.rows}
    export_rows: List[Dict[str, Any]] = []
    # Use the same parsing approach as _load_src_inventory for consistent keys.
    for rr in src_matrix0[1:]:
        if len(rr) < 6:
            continue
        w = as_str(rr[0]).strip()
        bc = as_str(rr[5]).strip()
        if w == "" or bc == "":
            continue
        export_rows.append(
            {
                "src_well": w,
                "barcode": bc,
                "volume_ul": raw_by_key.get((bc, w), as_number(rr[4]) or 0.0),
            }
        )
    after_dispense_path = out_dir / f"after_dispense_{inventory_xlsx.name}"
    _export_src_like_workbook(after_dispense_path, src_matrix0, export_rows, volume_col_idx0=4)
    print(f"Saved: {after_dispense_path}")

    # Plate:Works tables
    print("***** Plateworks Tables ******")
    print("***** Inventory")
    inventory_rows: List[List[Any]] = []
    for idx, bc in enumerate(inv.source_plate_barcodes, start=1):
        inventory_rows.append(make_inventory_row(bc, rack=1, pos=idx))
    for idx, dst in enumerate(dst_files, start=1):
        inventory_rows.append(make_inventory_row(dst.stem, rack=2, pos=idx))
    inventory_csv = "\n".join(",".join(as_str(x) for x in row) for row in inventory_rows) + "\n"
    inventory_path = out_dir / f"{picklist_base.name}_Inventory.csv"
    write_text(inventory_path, inventory_csv)
    print(inventory_csv.rstrip("\n"))
    print(f"Saving: {inventory_path}")

    print("***** Process_SRC")
    process_src = "\n".join(inv.source_plate_barcodes) + "\n"
    process_src_path = out_dir / f"{picklist_base.name}_Process_SRC.txt"
    write_text(process_src_path, process_src)
    print(process_src.rstrip("\n"))
    print(f"Saving: {process_src_path}")

    print("***** Process_DST")
    process_dst = "\n".join([p.stem for p in dst_files]) + "\n"
    process_dst_path = out_dir / f"{picklist_base.name}_Process_DST.txt"
    write_text(process_dst_path, process_dst)
    print(process_dst.rstrip("\n"))
    print(f"Saving: {process_dst_path}")

    print("***** These need to be manually added: ******")
    if manual_labels:
        for x in sorted(manual_labels):
            print(x)
    else:
        print("(none)")

    if warnings_flag[0]:
        print("Finished with warnings. Revise source file if needed!")
    else:
        print("Picklists passed all checks. Ready.")

    return RunResult(warnings=warnings_flag[0], outputs=[
        picklist_csv_path,
        inventory_path,
        process_src_path,
        process_dst_path,
        after_dispense_path,
    ])


def _parse_args(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Blue table picklist generator (PickliPy translation of bluetable.wls).",
    )
    p.add_argument(
        "picklist_base",
        help="Picklist base name (no extension). Can be a path; outputs go next to it.",
    )
    p.add_argument(
        "inventory_xlsx",
        help="Path to SRC inventory Excel file.",
    )
    p.add_argument(
        "folderdst",
        help="Folder containing destination blue-table .xlsx files (and optionally *_platesurvey.xml).",
    )
    p.add_argument(
        "--tables-out",
        default=None,
        help="Output folder for per-plate *_notebook_tables.xlsx (default: same as picklist outputs).",
    )
    p.add_argument(
        "--no-pause",
        action="store_true",
        help="Do not wait for Enter on exit (useful when running non-interactively).",
    )
    return p.parse_args(argv)


def main(argv: Optional[Sequence[str]] = None) -> None:
    args = _parse_args(argv)
    picklist_base = Path(args.picklist_base)
    inventory_xlsx = Path(args.inventory_xlsx)
    folderdst = Path(args.folderdst)
    tables_out = Path(args.tables_out) if args.tables_out else None

    def _run() -> None:
        run_bluetable(
            picklist_base=picklist_base,
            inventory_xlsx=inventory_xlsx,
            folderdst=folderdst,
            tables_out_dir=tables_out,
        )

    # Mirror wolframscript UX: pause unless --no-pause.
    raise SystemExit(run_with_user_facing_errors(_run, pause_on_exit=not args.no_pause))


if __name__ == "__main__":
    main()
