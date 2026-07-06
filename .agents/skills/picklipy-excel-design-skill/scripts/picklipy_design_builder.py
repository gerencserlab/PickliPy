#!/usr/bin/env python3
"""Build and validate Excel design files for PickliPy.Assay and PickliPy.Screen.

This helper creates design workbooks that the production PickliPy picklist
scripts can consume. It does not replace the production picklist generators.

The code intentionally writes the 384-well map geometry expected by the
Wolfram/Python generators: a 16 x 24 grid starting one cell down and one cell
right of labels such as "Plate Map:", "Concentrations Map:", and "Groups:".
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Optional, Sequence, Tuple

try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError as exc:  # pragma: no cover - explicit CLI message
    raise SystemExit(
        "This helper requires openpyxl. Install with: python -m pip install openpyxl"
    ) from exc

ROWS_384 = 16
COLS_384 = 24
ROW_NAMES_384 = [chr(ord("A") + i) for i in range(ROWS_384)]
LABEL_SEPARATORS_RE = re.compile(r"[+,;\n\r]+")
UNSAFE_LABEL_RE = re.compile(r"[+,;\n\r]")
WELL_RE = re.compile(r"^(?P<row>[A-Za-z]+)(?P<col>\d+)$")

HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
SECTION_FILL = PatternFill("solid", fgColor="E2F0D9")
MAP_FILL = PatternFill("solid", fgColor="F2F2F2")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ---------------------------------------------------------------------------
# Basic geometry and utility functions
# ---------------------------------------------------------------------------


def excel_col_to_number(col: str) -> int:
    col = col.strip().upper()
    if not re.fullmatch(r"[A-Z]+", col):
        raise ValueError(f"Invalid Excel column: {col!r}")
    n = 0
    for ch in col:
        n = n * 26 + ord(ch) - ord("A") + 1
    return n


def number_to_excel_col(n: int) -> str:
    if n < 1:
        raise ValueError("Column index must be >= 1")
    out: List[str] = []
    while n:
        n, rem = divmod(n - 1, 26)
        out.append(chr(ord("A") + rem))
    return "".join(reversed(out))


def parse_well(well: str) -> Tuple[int, int]:
    m = WELL_RE.match(str(well).strip())
    if not m:
        raise ValueError(f"Invalid well name: {well!r}")
    row = excel_col_to_number(m.group("row"))
    col = int(m.group("col"))
    if row < 1 or col < 1:
        raise ValueError(f"Invalid well name: {well!r}")
    return row, col


def well_name(row: int, col: int) -> str:
    return f"{number_to_excel_col(row)}{col}"


def all_wells(rows: int = ROWS_384, cols: int = COLS_384) -> List[str]:
    return [well_name(r, c) for r in range(1, rows + 1) for c in range(1, cols + 1)]


def is_blank(v: Any) -> bool:
    return v is None or (isinstance(v, str) and v.strip() == "")


def as_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and abs(v - round(v)) < 1e-9:
        return str(int(round(v)))
    return str(v)


def as_number(v: Any) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    if isinstance(v, str):
        s = v.strip()
        if s == "":
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def split_labels(text: Any) -> List[str]:
    if text is None:
        return []
    s = str(text)
    if s.strip() == "":
        return []
    return [p.strip() for p in LABEL_SEPARATORS_RE.split(s) if p.strip()]


def safe_label(label: str) -> str:
    label = str(label).strip()
    if label == "":
        raise ValueError("Label cannot be blank")
    if UNSAFE_LABEL_RE.search(label):
        raise ValueError(
            f"Unsafe PickliPy label {label!r}: labels may not contain comma, semicolon, plus, or newlines"
        )
    if any(ch.isspace() for ch in label):
        raise ValueError(f"Unsafe PickliPy label {label!r}: avoid spaces in labels")
    return label


def round_half_away_from_zero(x: float) -> int:
    return int(math.floor(x + 0.5)) if x >= 0 else int(math.ceil(x - 0.5))


def round_to_increment(x: float, inc: float = 2.5) -> float:
    if inc <= 0:
        raise ValueError("increment must be positive")
    return round_half_away_from_zero(x / inc) * inc


@dataclass
class TransferEstimate:
    desired_nl: float
    rounded_nl: float
    relative_error: Optional[float]
    warnings: List[str] = field(default_factory=list)


def estimate_transfer_volume(
    *,
    well_volume_ul: float,
    stock_concentration: float,
    final_concentration: float,
    source_plate_type: str = "LDV",
) -> TransferEstimate:
    if stock_concentration <= 0:
        raise ValueError("stock_concentration must be > 0")
    if well_volume_ul <= 0:
        raise ValueError("well_volume_ul must be > 0")
    if final_concentration < 0:
        raise ValueError("final_concentration cannot be negative")
    desired = 1000.0 * well_volume_ul * final_concentration / stock_concentration
    rounded = round_to_increment(desired, 2.5)
    rel = None if desired == 0 else abs(rounded - desired) / desired
    warnings: List[str] = []
    if rounded <= 0:
        warnings.append("invalid zero-volume transfer after 2.5 nL rounding")
    if rel is not None and rel > 0.10:
        warnings.append(f"large rounding error ({rel:.1%})")
    if source_plate_type.strip().upper() == "LDV" and rounded > 500:
        warnings.append("rounded transfer exceeds 500 nL LDV single-dispense limit")
    return TransferEstimate(desired, rounded, rel, warnings)


def blank_grid(value: str = "") -> List[List[str]]:
    return [[value for _ in range(COLS_384)] for _ in range(ROWS_384)]


def grid_from_well_map(well_map: Mapping[str, Any]) -> List[List[str]]:
    grid = blank_grid()
    for well, value in well_map.items():
        row, col = parse_well(str(well))
        if not (1 <= row <= ROWS_384 and 1 <= col <= COLS_384):
            raise ValueError(f"Well {well!r} is outside 384 map A1:P24")
        grid[row - 1][col - 1] = as_str(value)
    return grid


def normalize_grid(value: Any) -> List[List[str]]:
    if value is None:
        return blank_grid()
    if isinstance(value, Mapping):
        return grid_from_well_map(value)
    if isinstance(value, list):
        grid = blank_grid()
        for r, row in enumerate(value[:ROWS_384]):
            if not isinstance(row, list):
                raise ValueError("Grid rows must be lists")
            for c, cell in enumerate(row[:COLS_384]):
                grid[r][c] = as_str(cell)
        return grid
    raise ValueError("Expected plate map as well->value mapping or 16x24 list")


def labels_in_grid(grid: Sequence[Sequence[Any]]) -> List[str]:
    labels: List[str] = []
    for row in grid:
        for cell in row:
            labels.extend(split_labels(cell))
    seen: Dict[str, None] = {}
    for label in labels:
        seen.setdefault(label, None)
    return list(seen.keys())


def sorted_barcodes_from_csv(values: Sequence[str]) -> List[str]:
    return [b.strip() for b in ",".join(values).split(",") if b.strip()]


# ---------------------------------------------------------------------------
# Styling and worksheet writing
# ---------------------------------------------------------------------------


def set_common_sheet_style(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    ws.freeze_panes = "B2"
    widths = {
        "A": 20,
        "B": 22,
        "C": 18,
        "D": 22,
        "E": 18,
        "F": 18,
        "G": 16,
        "H": 18,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for col in range(9, 26):
        ws.column_dimensions[get_column_letter(col)].width = 10
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def style_range(ws: openpyxl.worksheet.worksheet.Worksheet, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_kv(ws: openpyxl.worksheet.worksheet.Worksheet, start_row: int, pairs: Sequence[Tuple[str, Any]]) -> int:
    r = start_row
    for key, value in pairs:
        ws.cell(r, 1).value = key
        ws.cell(r, 1).font = Font(bold=True)
        ws.cell(r, 1).fill = HEADER_FILL
        ws.cell(r, 2).value = value
        if isinstance(value, str) and value.startswith("@"):
            ws.cell(r, 2).number_format = "@"
        r += 1
    return r


def write_plate_grid(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    label_row: int,
    label: str,
    grid: Sequence[Sequence[Any]],
    *,
    section_fill: PatternFill = SECTION_FILL,
) -> None:
    """Write a PickliPy 16x24 map with visual row/column headers.

    The production generator reads only B:Y in the 16 rows immediately below
    the label row. The visual headers in A and on the label row are ignored.
    """
    ws.cell(label_row, 1).value = label
    ws.cell(label_row, 1).font = Font(bold=True)
    ws.cell(label_row, 1).fill = section_fill
    for c in range(1, COLS_384 + 1):
        cell = ws.cell(label_row, c + 1)
        cell.value = c
        cell.font = Font(bold=True)
        cell.fill = MAP_FILL
        cell.alignment = Alignment(horizontal="center")
    for r in range(1, ROWS_384 + 1):
        row_cell = ws.cell(label_row + r, 1)
        row_cell.value = number_to_excel_col(r)
        row_cell.font = Font(bold=True)
        row_cell.fill = MAP_FILL
        for c in range(1, COLS_384 + 1):
            value = grid[r - 1][c - 1]
            cell = ws.cell(label_row + r, c + 1)
            cell.value = value if value != "" else None
            if isinstance(value, str) and value.startswith("@"):
                cell.number_format = "@"
    style_range(ws, label_row, label_row + ROWS_384, 1, COLS_384 + 1)


def write_labels_table(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    start_row: int,
    rows: Sequence[Mapping[str, Any]],
) -> int:
    ws.cell(start_row, 1).value = "Labels"
    ws.cell(start_row, 2).value = "Compounds"
    ws.cell(start_row, 3).value = "Concentrations"
    for c in range(1, 4):
        ws.cell(start_row, c).font = Font(bold=True)
        ws.cell(start_row, c).fill = SECTION_FILL
        ws.cell(start_row, c).border = BORDER
    r = start_row + 1
    for item in rows:
        label = safe_label(as_str(item.get("label", item.get("Labels", ""))))
        compound = as_str(item.get("compound", item.get("Compounds", ""))).strip()
        concentration = item.get("concentration", item.get("Concentrations", ""))
        if compound == "":
            raise ValueError(f"Label {label!r} has no compound mapping")
        ws.cell(r, 1).value = label
        ws.cell(r, 2).value = compound
        ws.cell(r, 3).value = concentration
        for c in range(1, 4):
            ws.cell(r, c).border = BORDER
            if isinstance(ws.cell(r, c).value, str) and str(ws.cell(r, c).value).startswith("@"):
                ws.cell(r, c).number_format = "@"
        r += 1
    return r


def save_workbook(wb: Workbook, output: str | Path) -> Path:
    out = Path(output)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


# ---------------------------------------------------------------------------
# Assay workbook creation
# ---------------------------------------------------------------------------


DEFAULT_ASSAY_HEADER = {
    "Well volume (ul):": 50,
    "Inventory_SRC:": "Inventory_SRC",
    "Inventory_DST:": "Inventory_DST",
    "Inventory_SRC_Rack#:": 1,
    "Inventory_DST_Rack#:": 2,
    "Process_SRC:": "Process_SRC",
    "Process_DST:": "Process_DST",
}


@dataclass
class AssaySourceRow:
    source_well: str
    compound_name: str
    stock_concentration: float
    unit: str = "uM"
    volume_ul: float = 30.0
    plate_barcode: str = "PlateSRC"


@dataclass
class AssayAddition:
    picklist_name: str
    barcode_src: str
    barcode_dst: str | Sequence[str]
    plate_map: Any
    labels: Sequence[Mapping[str, Any]]
    concentration_map: Any = None
    well_volume_ul: Optional[float] = None


class ValidationError(RuntimeError):
    pass


def make_assay_workbook(
    *,
    output: str | Path,
    source_rows: Sequence[Mapping[str, Any] | AssaySourceRow],
    additions: Sequence[Mapping[str, Any] | AssayAddition],
    header: Optional[Mapping[str, Any]] = None,
    min_source_volume_ul: float = 15.0,
) -> Path:
    wb = Workbook()
    ws_src = wb.active
    ws_src.title = "SRC"
    ws_dst = wb.create_sheet("DST")

    # SRC sheet
    src_headers = [
        "Source Well",
        "Compound Name",
        "Stock concentration",
        "Unit (same unit as in DST)",
        "Volume in source plate",
        "Plate barcode",
    ]
    for c, name in enumerate(src_headers, 1):
        ws_src.cell(1, c).value = name
        ws_src.cell(1, c).font = Font(bold=True)
        ws_src.cell(1, c).fill = HEADER_FILL
        ws_src.cell(1, c).border = BORDER
    ws_src.cell(1, 7).value = min_source_volume_ul
    ws_src.cell(1, 8).value = "<-Min volume"
    ws_src.cell(1, 7).fill = WARN_FILL
    ws_src.cell(1, 8).fill = WARN_FILL

    for r, row in enumerate(source_rows, 2):
        if isinstance(row, AssaySourceRow):
            d = row.__dict__
        else:
            d = dict(row)
        values = [
            d.get("source_well", d.get("Source Well")),
            d.get("compound_name", d.get("Compound Name")),
            d.get("stock_concentration", d.get("Stock concentration")),
            d.get("unit", d.get("Unit (same unit as in DST)", "uM")),
            d.get("volume_ul", d.get("Volume in source plate", 30.0)),
            d.get("plate_barcode", d.get("Plate barcode", "PlateSRC")),
        ]
        for c, value in enumerate(values, 1):
            ws_src.cell(r, c).value = value
            ws_src.cell(r, c).border = BORDER
    set_common_sheet_style(ws_src)

    # DST header
    h = dict(DEFAULT_ASSAY_HEADER)
    if header:
        h.update(header)
    write_kv(ws_dst, 1, list(h.items()))

    current_row = 10
    for idx, raw_addition in enumerate(additions, 1):
        add = raw_addition if isinstance(raw_addition, AssayAddition) else AssayAddition(**dict(raw_addition))
        plate_grid = normalize_grid(add.plate_map)
        for label in labels_in_grid(plate_grid):
            safe_label(label)
        write_plate_grid(ws_dst, current_row, "Plate Map:", plate_grid)
        current_row += ROWS_384 + 2

        if add.concentration_map is not None:
            conc_grid = normalize_grid(add.concentration_map)
            write_plate_grid(ws_dst, current_row, "Concentrations Map:", conc_grid)
            current_row += ROWS_384 + 2

        if add.well_volume_ul is not None:
            current_row = write_kv(ws_dst, current_row, [("Well volume (ul):", add.well_volume_ul)])
        dst_barcodes = add.barcode_dst if isinstance(add.barcode_dst, str) else ",".join(add.barcode_dst)
        current_row = write_kv(
            ws_dst,
            current_row,
            [
                ("Picklist Name:", add.picklist_name),
                ("Barcode_SRC:", add.barcode_src),
                ("Barcode_DST:", dst_barcodes),
            ],
        )
        current_row = write_labels_table(ws_dst, current_row, add.labels)
        current_row += 2

    set_common_sheet_style(ws_dst)
    path = save_workbook(wb, output)
    return path


# ---------------------------------------------------------------------------
# Screen workbook creation
# ---------------------------------------------------------------------------


DEFAULT_SCREEN_HEADER = {
    "Well volume (ul):": 50,
    "Inventory_SRC:": "Inventory_SRC",
    "Inventory_DST:": "Inventory_DST",
    "Inventory_SRC_Rack#:": 1,
    "Inventory_DST_Rack#:": 2,
    "Process_SRC:": "Process_SRC",
    "Process_DST:": "Process_DST",
    "Loop Counts_SRC:": "Loop_SRC",
    "Loop Counts_DST:": "Loop_DST",
}


@dataclass
class ScreenControl:
    source_well: str
    compound_name: str
    stock_concentration: float
    identifier: str
    volume_ul: float = 10.0
    plate_barcode: str = "*"


@dataclass
class ScreenLayout:
    plate_map: List[List[str]]
    labels: List[Dict[str, Any]]
    groups: Optional[List[List[str]]] = None


def candidate_wells_for_layout(
    *,
    reserve_edges: bool = True,
    rows: int = ROWS_384,
    cols: int = COLS_384,
    order: str = "row-major",
) -> List[str]:
    wells: List[str] = []
    for r in range(1, rows + 1):
        columns = list(range(1, cols + 1))
        if order == "serpentine" and r % 2 == 0:
            columns.reverse()
        for c in columns:
            if reserve_edges and (r in (1, rows) or c in (1, cols)):
                continue
            wells.append(well_name(r, c))
    return wells


def make_screen_slot_layout(
    *,
    slots: int = 77,
    replicates: int = 1,
    concentrations: Optional[Sequence[Any]] = None,
    reserve_edges: bool = True,
    control_label: Optional[str] = "veh",
    positive_label: Optional[str] = "pos",
    fill_label: str = "fill",
    order: str = "row-major",
    group_mode: str = "all",
) -> ScreenLayout:
    if slots < 1:
        raise ValueError("slots must be >= 1")
    if replicates < 1:
        raise ValueError("replicates must be >= 1")
    concs = list(concentrations) if concentrations else [10]
    plate_map = blank_grid()
    labels: List[Dict[str, Any]] = []

    # Controls on edge wells if requested.
    if reserve_edges and control_label:
        for r in range(1, ROWS_384 + 1):
            for c in range(1, COLS_384 + 1):
                if r in (1, ROWS_384) or c in (1, COLS_384):
                    plate_map[r - 1][c - 1] = control_label
        labels.append({"label": control_label, "compound": control_label, "concentration": 0})
        if positive_label:
            # Place positives in four corners, overriding vehicle edge controls.
            for w in ["A1", "A24", "P1", "P24"]:
                r, c = parse_well(w)
                plate_map[r - 1][c - 1] = positive_label
            labels.append({"label": positive_label, "compound": positive_label, "concentration": 1})

    wells = candidate_wells_for_layout(reserve_edges=reserve_edges, order=order)
    required = slots * replicates * len(concs)
    if required > len(wells):
        raise ValueError(f"Layout needs {required} treatment wells but only {len(wells)} are available")

    idx = 0
    for slot in range(1, slots + 1):
        compound_item = f"#{slot}"
        for dose_index, conc in enumerate(concs, 1):
            label = compound_item if len(concs) == 1 else f"#{slot}_d{dose_index}"
            safe_label(label)
            labels.append({"label": label, "compound": compound_item, "concentration": conc})
            for _rep in range(replicates):
                r, c = parse_well(wells[idx])
                plate_map[r - 1][c - 1] = label
                idx += 1

    # Final row is the blacklist fill label.
    if fill_label:
        labels.append({"label": fill_label, "compound": control_label or fill_label, "concentration": 0})

    groups = make_group_grid(plate_map, mode=group_mode, controls={control_label, positive_label, fill_label, "", None})
    return ScreenLayout(plate_map=plate_map, labels=labels, groups=groups)


def make_group_grid(
    plate_map: Sequence[Sequence[Any]],
    *,
    mode: str = "all",
    controls: Iterable[Any] = ("veh", "pos", "neg", "blank", "fill", ""),
) -> List[List[str]]:
    controls_set = {as_str(x) for x in controls if x is not None}
    groups = blank_grid()
    for r in range(1, ROWS_384 + 1):
        for c in range(1, COLS_384 + 1):
            v = as_str(plate_map[r - 1][c - 1]).strip()
            if v == "" or v in controls_set:
                continue
            if mode == "none":
                groups[r - 1][c - 1] = ""
            elif mode == "halves":
                groups[r - 1][c - 1] = "left" if c <= COLS_384 // 2 else "right"
            elif mode == "quadrants":
                top = "top" if r <= ROWS_384 // 2 else "bottom"
                side = "left" if c <= COLS_384 // 2 else "right"
                groups[r - 1][c - 1] = f"{top}_{side}"
            else:
                groups[r - 1][c - 1] = "1"
    return groups


def make_screen_workbook(
    *,
    output: str | Path,
    library_rows: Sequence[Mapping[str, Any]],
    library_headers: Optional[Sequence[str]] = None,
    controls: Optional[Sequence[Mapping[str, Any] | ScreenControl]] = None,
    layout: Optional[ScreenLayout | Mapping[str, Any]] = None,
    header: Optional[Mapping[str, Any]] = None,
    picklist_name: str = "Screen",
    dst_barcodes: Sequence[str] = ("Plate1", "Plate2", "Plate3"),
    min_source_volume_ul: float = 2.5,
    src_reference_row: Optional[Mapping[str, Any]] = None,
    blacklist: Optional[Mapping[str, Sequence[str]]] = None,
    include_blacklist: bool = True,
    sort_library_by_barcode: bool = True,
) -> Path:
    wb = Workbook()
    ws_src = wb.active
    ws_src.title = "SRC"
    ws_dst = wb.create_sheet("DST")
    ws_lib = wb.create_sheet("LIB")

    # Default library header scheme: SRC references A-D.
    if library_headers is None:
        library_headers = ["Source Well", "Identifier", "Stock concentration", "Plate barcode", "Final concentration"]

    # Normalize and optionally sort library rows.
    lib_rows = [dict(r) for r in library_rows]
    if sort_library_by_barcode:
        # Preserve within-barcode order using stable sort.
        barcode_key = find_header_key(lib_rows, ["Plate barcode", "Barcode", "Source Plate Barcode"])
        if barcode_key:
            lib_rows = sorted(lib_rows, key=lambda row: as_str(row.get(barcode_key, "")))

    # SRC sheet.
    src_headers = [
        "Source Well",
        "Compound Name",
        "Stock concentration",
        "Identifier for merging",
        "Volume in source plate",
        "Plate barcode",
    ]
    for c, name in enumerate(src_headers, 1):
        ws_src.cell(1, c).value = name
        ws_src.cell(1, c).font = Font(bold=True)
        ws_src.cell(1, c).fill = HEADER_FILL
        ws_src.cell(1, c).border = BORDER
    ws_src.cell(1, 7).value = min_source_volume_ul
    ws_src.cell(1, 8).value = "<-Min volume"
    ws_src.cell(1, 7).fill = WARN_FILL
    ws_src.cell(1, 8).fill = WARN_FILL

    r = 2
    for raw in controls or []:
        ctrl = raw if isinstance(raw, ScreenControl) else ScreenControl(**dict(raw))
        values = [
            ctrl.source_well,
            ctrl.compound_name,
            ctrl.stock_concentration,
            ctrl.identifier,
            ctrl.volume_ul,
            ctrl.plate_barcode,
        ]
        for c, value in enumerate(values, 1):
            ws_src.cell(r, c).value = value
            ws_src.cell(r, c).border = BORDER
        r += 1

    ref = dict(
        source_well="@A",
        compound_name="*",
        stock_concentration="@C",
        identifier="@B",
        volume_ul=5,
        plate_barcode="@D",
    )
    if src_reference_row:
        ref.update(src_reference_row)
    ref_values = [
        ref["source_well"],
        ref["compound_name"],
        ref["stock_concentration"],
        ref["identifier"],
        ref["volume_ul"],
        ref["plate_barcode"],
    ]
    for c, value in enumerate(ref_values, 1):
        cell = ws_src.cell(r, c)
        cell.value = value
        if isinstance(value, str) and value.startswith("@"):
            cell.number_format = "@"
        cell.border = BORDER

    # DST header.
    h = dict(DEFAULT_SCREEN_HEADER)
    if header:
        h.update(header)
    write_kv(ws_dst, 1, list(h.items()))

    if layout is None:
        layout_obj = make_screen_slot_layout()
    elif isinstance(layout, ScreenLayout):
        layout_obj = layout
    else:
        layout_obj = ScreenLayout(
            plate_map=normalize_grid(layout.get("plate_map")),
            labels=list(layout.get("labels", [])),
            groups=normalize_grid(layout.get("groups")) if layout.get("groups") is not None else None,
        )

    current_row = 14
    write_plate_grid(ws_dst, current_row, "Plate Map:", layout_obj.plate_map)
    current_row += ROWS_384 + 2
    current_row = write_kv(
        ws_dst,
        current_row,
        [
            ("Picklist Name:", picklist_name),
            ("Barcode_DST:", ",".join(dst_barcodes)),
        ],
    )
    write_labels_table(ws_dst, current_row, layout_obj.labels)

    # LIB sheet.
    for c, name in enumerate(library_headers, 1):
        ws_lib.cell(1, c).value = name
        ws_lib.cell(1, c).font = Font(bold=True)
        ws_lib.cell(1, c).fill = HEADER_FILL
        ws_lib.cell(1, c).border = BORDER
    for r, row in enumerate(lib_rows, 2):
        for c, header_name in enumerate(library_headers, 1):
            ws_lib.cell(r, c).value = row.get(header_name, "")
            ws_lib.cell(r, c).border = BORDER

    # Optional blacklist sheet.
    if include_blacklist:
        ws_blk = wb.create_sheet("DST_Blacklist")
        groups = layout_obj.groups or make_group_grid(layout_obj.plate_map)
        write_plate_grid(ws_blk, 1, "Groups:", groups)
        start = ROWS_384 + 4
        ws_blk.cell(start, 1).value = "Barcode_DST"
        ws_blk.cell(start, 2).value = "Well"
        ws_blk.cell(start, 1).font = Font(bold=True)
        ws_blk.cell(start, 2).font = Font(bold=True)
        ws_blk.cell(start, 1).fill = SECTION_FILL
        ws_blk.cell(start, 2).fill = SECTION_FILL
        for i, barcode in enumerate(dst_barcodes, start + 1):
            ws_blk.cell(i, 1).value = barcode
            wells = blacklist.get(barcode, []) if blacklist else []
            ws_blk.cell(i, 2).value = ",".join(wells)
        set_common_sheet_style(ws_blk)

    for ws in wb.worksheets:
        set_common_sheet_style(ws)

    path = save_workbook(wb, output)
    return path


def find_header_key(rows: Sequence[Mapping[str, Any]], candidates: Sequence[str]) -> Optional[str]:
    if not rows:
        return None
    keys = list(rows[0].keys())
    lower_to_key = {k.lower(): k for k in keys}
    for cand in candidates:
        if cand.lower() in lower_to_key:
            return lower_to_key[cand.lower()]
    return None


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------


def read_grid(ws: openpyxl.worksheet.worksheet.Worksheet, label_row: int) -> List[List[str]]:
    return [
        [as_str(ws.cell(label_row + r, 1 + c).value).strip() for c in range(1, COLS_384 + 1)]
        for r in range(1, ROWS_384 + 1)
    ]


def find_rows_with_label(ws: openpyxl.worksheet.worksheet.Worksheet, label: str) -> List[int]:
    rows: List[int] = []
    for r in range(1, ws.max_row + 1):
        if as_str(ws.cell(r, 1).value) == label:
            rows.append(r)
    return rows


def read_labels_table(ws: openpyxl.worksheet.worksheet.Worksheet, labels_row: int) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    r = labels_row + 1
    while r <= ws.max_row:
        label = as_str(ws.cell(r, 1).value).strip()
        if label == "" or label in {"Plate Map:", "Groups:"}:
            break
        rows.append(
            {
                "label": label,
                "compound": as_str(ws.cell(r, 2).value).strip(),
                "concentration": ws.cell(r, 3).value,
            }
        )
        r += 1
    return rows


def validate_assay(path: str | Path) -> List[str]:
    wb = load_workbook(path, data_only=True)
    if len(wb.worksheets) < 2:
        raise ValidationError("Assay workbook requires at least SRC and DST worksheets")
    ws_src, ws_dst = wb.worksheets[0], wb.worksheets[1]
    warnings: List[str] = []

    # Source uniqueness and compound stock consistency.
    seen_wells: Dict[Tuple[str, str], int] = {}
    compound_stock: Dict[Tuple[str, str], float] = {}
    src_compounds_by_plate: Dict[str, set[str]] = {}
    for r in range(2, ws_src.max_row + 1):
        well = as_str(ws_src.cell(r, 1).value).strip()
        comp = as_str(ws_src.cell(r, 2).value).strip()
        stock = as_number(ws_src.cell(r, 3).value)
        barcode = as_str(ws_src.cell(r, 6).value).strip()
        if not well and not comp and not barcode:
            continue
        if well:
            parse_well(well)
        key = (barcode, well)
        if key in seen_wells:
            raise ValidationError(f"Duplicate source well {well} in source plate {barcode}")
        seen_wells[key] = r
        if comp and barcode:
            src_compounds_by_plate.setdefault(barcode, set()).add(comp)
        if comp and stock is not None:
            ckey = (barcode, comp)
            if ckey in compound_stock and abs(compound_stock[ckey] - stock) > 1e-9:
                raise ValidationError(f"Compound {comp} has inconsistent stock concentration in source plate {barcode}")
            compound_stock[ckey] = stock

    plate_map_rows = find_rows_with_label(ws_dst, "Plate Map:")
    if not plate_map_rows:
        raise ValidationError("No Plate Map: section found in DST")
    labels_rows = find_rows_with_label(ws_dst, "Labels")
    if not labels_rows:
        raise ValidationError("No Labels table found in DST")

    for pm_row in plate_map_rows:
        next_pm = min([r for r in plate_map_rows if r > pm_row], default=ws_dst.max_row + 1)
        labels_row = min([r for r in labels_rows if pm_row < r < next_pm], default=None)
        if labels_row is None:
            raise ValidationError(f"Plate Map at row {pm_row} has no Labels table")
        grid = read_grid(ws_dst, pm_row)
        used = set(labels_in_grid(grid))
        rows = read_labels_table(ws_dst, labels_row)
        defined = {row["label"] for row in rows}
        missing = sorted(used - defined)
        if missing:
            raise ValidationError(f"Plate Map at row {pm_row} has labels missing from Labels table: {missing}")
        for label in used:
            safe_label(label)

        # Find Barcode_SRC between plate-map and labels table.
        barcode_src = None
        for r in range(pm_row + 1, labels_row):
            if as_str(ws_dst.cell(r, 1).value) == "Barcode_SRC:":
                barcode_src = as_str(ws_dst.cell(r, 2).value).strip()
                break
        if barcode_src:
            for row in rows:
                comp = row["compound"]
                if comp and comp not in src_compounds_by_plate.get(barcode_src, set()):
                    raise ValidationError(
                        f"Compound {comp!r} in Labels table for source {barcode_src!r} is absent from SRC"
                    )
        else:
            warnings.append(f"Plate Map at row {pm_row}: Barcode_SRC: not found")

    return warnings


def validate_screen(path: str | Path) -> List[str]:
    wb = load_workbook(path, data_only=True)
    if len(wb.worksheets) < 3:
        raise ValidationError("Screen workbook requires SRC, DST, and LIB worksheets")
    ws_src, ws_dst, ws_lib = wb.worksheets[0], wb.worksheets[1], wb.worksheets[2]
    warnings: List[str] = []

    plate_map_rows = find_rows_with_label(ws_dst, "Plate Map:")
    if len(plate_map_rows) != 1:
        raise ValidationError(f"Screen workbook should have exactly one Plate Map:, found {len(plate_map_rows)}")
    labels_rows = find_rows_with_label(ws_dst, "Labels")
    if len(labels_rows) != 1:
        raise ValidationError(f"Screen workbook should have exactly one Labels table, found {len(labels_rows)}")

    grid = read_grid(ws_dst, plate_map_rows[0])
    used = set(labels_in_grid(grid))
    label_rows = read_labels_table(ws_dst, labels_rows[0])
    defined = {row["label"] for row in label_rows}
    missing = sorted(used - defined)
    if missing:
        raise ValidationError(f"Screen Plate Map has labels missing from Labels table: {missing}")

    # Barcode list.
    barcode_dst = None
    for r in range(1, ws_dst.max_row + 1):
        if as_str(ws_dst.cell(r, 1).value) == "Barcode_DST:":
            barcode_dst = as_str(ws_dst.cell(r, 2).value)
            break
    if not barcode_dst:
        raise ValidationError("Screen DST is missing Barcode_DST:")
    barcodes = [b.strip() for b in barcode_dst.split(",") if b.strip()]
    if not barcodes:
        raise ValidationError("Screen Barcode_DST: has no barcodes")

    # Check LIB source barcode order using likely header names.
    headers = [as_str(ws_lib.cell(1, c).value).strip() for c in range(1, ws_lib.max_column + 1)]
    header_index = {h.lower(): i + 1 for i, h in enumerate(headers)}
    barcode_col = None
    for cand in ["plate barcode", "barcode", "source plate barcode"]:
        if cand in header_index:
            barcode_col = header_index[cand]
            break
    if barcode_col is not None:
        seen_order: List[str] = []
        seen_set: set[str] = set()
        last = None
        for r in range(2, ws_lib.max_row + 1):
            first_col_value = as_str(ws_lib.cell(r, 1).value).strip()
            if first_col_value == "":
                continue
            barcode = as_str(ws_lib.cell(r, barcode_col).value).strip()
            if barcode != last:
                if barcode in seen_set:
                    raise ValidationError(
                        f"LIB source plate barcode {barcode!r} is revisited after another barcode. Sort LIB by source barcode."
                    )
                seen_set.add(barcode)
                seen_order.append(barcode)
                last = barcode
    else:
        warnings.append("Could not identify LIB source plate barcode column for sorting check")

    # Check SRC @ references refer to existing LIB columns.
    max_col = ws_lib.max_column
    for r in range(2, ws_src.max_row + 1):
        for c in range(1, 7):
            v = as_str(ws_src.cell(r, c).value).strip()
            if v.startswith("@"):
                col_letters = v[1:].strip()
                if not re.fullmatch(r"[A-Za-z]+", col_letters):
                    raise ValidationError(f"Invalid @ reference {v!r} in SRC row {r}")
                if excel_col_to_number(col_letters) > max_col:
                    raise ValidationError(f"SRC reference {v!r} points beyond LIB max column {max_col}")

    return warnings


def validate_workbook(path: str | Path, mode: str) -> None:
    if mode == "assay":
        warnings = validate_assay(path)
    elif mode == "screen":
        warnings = validate_screen(path)
    else:
        raise ValueError("mode must be assay or screen")
    print(f"Validation passed for {mode}: {path}")
    for warning in warnings:
        print(f"Warning: {warning}")


# ---------------------------------------------------------------------------
# CSV/JSON loading and shortlisting
# ---------------------------------------------------------------------------


def read_csv_dicts(path: str | Path) -> List[Dict[str, Any]]:
    with Path(path).open("r", newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def write_csv_rows(path: str | Path, rows: Sequence[Sequence[Any]]) -> Path:
    out = Path(path)
    out.parent.mkdir(parents=True, exist_ok=True)
    with out.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerows(rows)
    return out


def load_json(path: str | Path) -> Dict[str, Any]:
    with Path(path).open("r", encoding="utf-8") as f:
        return json.load(f)


def shortlist_rows(rows: Sequence[Mapping[str, Any]], filters: Optional[Mapping[str, Any]]) -> List[Dict[str, Any]]:
    if not filters:
        return [dict(r) for r in rows]
    out: List[Dict[str, Any]] = []
    for row in rows:
        keep = True
        for key, condition in filters.items():
            value = row.get(key)
            if isinstance(condition, list):
                keep = keep and value in condition
            elif isinstance(condition, dict):
                num = as_number(value)
                if num is None:
                    keep = False
                if "min" in condition and num is not None and num < float(condition["min"]):
                    keep = False
                if "max" in condition and num is not None and num > float(condition["max"]):
                    keep = False
                if "equals" in condition and value != condition["equals"]:
                    keep = False
            else:
                keep = keep and value == condition
        if keep:
            out.append(dict(row))
    return out


# ---------------------------------------------------------------------------
# Blacklist generation
# ---------------------------------------------------------------------------


def build_blacklist_from_csv(
    *,
    input_csv: str | Path,
    output_csv: str | Path,
    barcode_col: str,
    well_col: str,
    metric_col: str,
    min_value: Optional[float] = None,
    max_value: Optional[float] = None,
) -> Path:
    rows = read_csv_dicts(input_csv)
    grouped: Dict[str, List[str]] = {}
    for row in rows:
        barcode = as_str(row.get(barcode_col, "")).strip()
        well = as_str(row.get(well_col, "")).strip()
        metric = as_number(row.get(metric_col))
        if not barcode or not well:
            continue
        reject = metric is None
        if min_value is not None and metric is not None and metric < min_value:
            reject = True
        if max_value is not None and metric is not None and metric > max_value:
            reject = True
        if reject:
            parse_well(well)
            grouped.setdefault(barcode, []).append(well)
    out_rows = [["Barcode_DST", "Well"]]
    for barcode in sorted(grouped):
        out_rows.append([barcode, ",".join(grouped[barcode])])
    return write_csv_rows(output_csv, out_rows)


# ---------------------------------------------------------------------------
# Demo builders and CLI
# ---------------------------------------------------------------------------


def assay_demo(output: str | Path) -> Path:
    dose_wells = {}
    conc_map = {}
    concentrations = [0.03, 0.1, 0.3, 1, 3, 10, 30, 100]
    for r in range(3, 11):  # C-J replicates
        for i, conc in enumerate(concentrations, start=4):  # columns 4-11
            dose_wells[well_name(r, i)] = "FCCP"
            conc_map[well_name(r, i)] = conc
    # A second combinatorial section to demonstrate merging with same picklist.
    combo_map = {
        "L4": "OLIGO+ROT",
        "L5": "OLIGO+ROT",
        "L6": "OLIGO+FCCP",
        "L7": "OLIGO+FCCP",
    }
    source_rows = [
        AssaySourceRow("A1", "FCCP", 2000, "uM", 30, "ToolSRC"),
        AssaySourceRow("A2", "OLIGO", 1000, "uM", 30, "ToolSRC"),
        AssaySourceRow("A3", "ROT", 1000, "uM", 30, "ToolSRC"),
    ]
    additions = [
        AssayAddition(
            picklist_name="Assay_Demo",
            barcode_src="ToolSRC",
            barcode_dst=["CellPlate1", "CellPlate2"],
            plate_map=dose_wells,
            concentration_map=conc_map,
            labels=[{"label": "FCCP", "compound": "FCCP", "concentration": 0}],
        ),
        AssayAddition(
            picklist_name="Assay_Demo",
            barcode_src="ToolSRC",
            barcode_dst=["CellPlate1", "CellPlate2"],
            plate_map=combo_map,
            labels=[
                {"label": "OLIGO", "compound": "OLIGO", "concentration": 1},
                {"label": "ROT", "compound": "ROT", "concentration": 1},
                {"label": "FCCP", "compound": "FCCP", "concentration": 1},
            ],
        ),
    ]
    return make_assay_workbook(output=output, source_rows=source_rows, additions=additions)


def screen_demo(output: str | Path, *, slots: int, dst_barcodes: Sequence[str]) -> Path:
    layout = make_screen_slot_layout(slots=slots, replicates=1, concentrations=[10], reserve_edges=True)
    library_rows: List[Dict[str, Any]] = []
    for i in range(1, slots + 1):
        plate = "LibPlate1" if i <= max(1, slots // 2) else "LibPlate2"
        library_rows.append(
            {
                "Source Well": well_name(((i - 1) % ROWS_384) + 1, ((i - 1) // ROWS_384) + 1),
                "Identifier": f"Cmpd_{i:03d}",
                "Stock concentration": 10000,
                "Plate barcode": plate,
                "Final concentration": 10,
            }
        )
    controls = [
        ScreenControl("A1", "veh", 10000, "DMSO", 10, "*"),
        ScreenControl("A2", "pos", 10000, "PositiveControl", 10, "*"),
    ]
    return make_screen_workbook(
        output=output,
        library_rows=library_rows,
        controls=controls,
        layout=layout,
        picklist_name="Screen_Demo",
        dst_barcodes=dst_barcodes,
        include_blacklist=True,
    )


def cmd_assay_from_spec(args: argparse.Namespace) -> None:
    spec = load_json(args.spec)
    out = make_assay_workbook(
        output=args.output,
        source_rows=spec["source_rows"],
        additions=spec["additions"],
        header=spec.get("header"),
        min_source_volume_ul=spec.get("min_source_volume_ul", 15.0),
    )
    print(out)
    validate_workbook(out, "assay")


def cmd_screen_from_spec(args: argparse.Namespace) -> None:
    spec = load_json(args.spec)
    library_rows = spec.get("library_rows", [])
    if "library_csv" in spec:
        library_rows = read_csv_dicts(spec["library_csv"])
    library_rows = shortlist_rows(library_rows, spec.get("shortlist_filters"))
    layout_spec = spec.get("layout")
    if layout_spec and layout_spec.get("type") == "slot_layout":
        layout = make_screen_slot_layout(
            slots=layout_spec.get("slots", 77),
            replicates=layout_spec.get("replicates", 1),
            concentrations=layout_spec.get("concentrations", [10]),
            reserve_edges=layout_spec.get("reserve_edges", True),
            control_label=layout_spec.get("control_label", "veh"),
            positive_label=layout_spec.get("positive_label", "pos"),
            fill_label=layout_spec.get("fill_label", "fill"),
            order=layout_spec.get("order", "row-major"),
            group_mode=layout_spec.get("group_mode", "all"),
        )
    elif layout_spec:
        layout = ScreenLayout(
            plate_map=normalize_grid(layout_spec.get("plate_map")),
            labels=layout_spec.get("labels", []),
            groups=normalize_grid(layout_spec.get("groups")) if layout_spec.get("groups") is not None else None,
        )
    else:
        layout = make_screen_slot_layout()
    out = make_screen_workbook(
        output=args.output,
        library_rows=library_rows,
        library_headers=spec.get("library_headers"),
        controls=spec.get("controls", []),
        layout=layout,
        header=spec.get("header"),
        picklist_name=spec.get("picklist_name", "Screen"),
        dst_barcodes=spec.get("dst_barcodes", ["Plate1", "Plate2", "Plate3"]),
        min_source_volume_ul=spec.get("min_source_volume_ul", 2.5),
        src_reference_row=spec.get("src_reference_row"),
        blacklist=spec.get("blacklist"),
        include_blacklist=spec.get("include_blacklist", True),
        sort_library_by_barcode=spec.get("sort_library_by_barcode", True),
    )
    print(out)
    validate_workbook(out, "screen")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Create and validate PickliPy Excel design workbooks")
    sub = parser.add_subparsers(dest="cmd", required=True)

    p = sub.add_parser("assay-demo", help="Create an example Assay workbook")
    p.add_argument("--output", required=True)
    p.set_defaults(func=lambda args: print(assay_demo(args.output)))

    p = sub.add_parser("screen-demo", help="Create an example Screen workbook")
    p.add_argument("--output", required=True)
    p.add_argument("--slots", type=int, default=77)
    p.add_argument("--dst-barcodes", default="Plate1,Plate2,Plate3")
    p.set_defaults(
        func=lambda args: print(
            screen_demo(args.output, slots=args.slots, dst_barcodes=[b.strip() for b in args.dst_barcodes.split(",") if b.strip()])
        )
    )

    p = sub.add_parser("assay-from-spec", help="Create an Assay workbook from a JSON spec")
    p.add_argument("spec")
    p.add_argument("--output", required=True)
    p.set_defaults(func=cmd_assay_from_spec)

    p = sub.add_parser("screen-from-spec", help="Create a Screen workbook from a JSON spec")
    p.add_argument("spec")
    p.add_argument("--output", required=True)
    p.set_defaults(func=cmd_screen_from_spec)

    p = sub.add_parser("validate", help="Validate an existing PickliPy design workbook")
    p.add_argument("workbook")
    p.add_argument("--mode", required=True, choices=["assay", "screen"])
    p.set_defaults(func=lambda args: validate_workbook(args.workbook, args.mode))

    p = sub.add_parser("estimate-volume", help="Estimate Echo transfer volume and rounding")
    p.add_argument("--well-volume-ul", type=float, required=True)
    p.add_argument("--stock", type=float, required=True, help="Source stock concentration")
    p.add_argument("--final", type=float, required=True, help="Final destination concentration in same unit as stock")
    p.add_argument("--source-plate-type", default="LDV")

    def _estimate(args: argparse.Namespace) -> None:
        est = estimate_transfer_volume(
            well_volume_ul=args.well_volume_ul,
            stock_concentration=args.stock,
            final_concentration=args.final,
            source_plate_type=args.source_plate_type,
        )
        rel = "n/a" if est.relative_error is None else f"{est.relative_error:.3%}"
        print(f"desired_nL={est.desired_nl:.6g}")
        print(f"rounded_nL={est.rounded_nl:.6g}")
        print(f"relative_error={rel}")
        for w in est.warnings:
            print(f"Warning: {w}")

    p.set_defaults(func=_estimate)

    p = sub.add_parser("blacklist-from-csv", help="Create Barcode_DST/Well blacklist CSV from a per-well metric table")
    p.add_argument("input_csv")
    p.add_argument("--output", required=True)
    p.add_argument("--barcode-col", required=True)
    p.add_argument("--well-col", required=True)
    p.add_argument("--metric-col", required=True)
    p.add_argument("--min-value", type=float)
    p.add_argument("--max-value", type=float)
    p.set_defaults(
        func=lambda args: print(
            build_blacklist_from_csv(
                input_csv=args.input_csv,
                output_csv=args.output,
                barcode_col=args.barcode_col,
                well_col=args.well_col,
                metric_col=args.metric_col,
                min_value=args.min_value,
                max_value=args.max_value,
            )
        )
    )

    return parser


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    try:
        args.func(args)
        return 0
    except (ValidationError, ValueError, KeyError) as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 2


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(main())
